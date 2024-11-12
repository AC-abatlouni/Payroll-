import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
import sys
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, List

# Define threshold tables
hvac_thresholds = {
    0: [7000, 8000, 9000, 10000],
    10: [7500, 8500, 9500, 11000],
    20: [8000, 9000, 10000, 11000],
    30: [9000, 10000, 11000, 12000],
    40: [10000, 11000, 12000, 13000],
    50: [12000, 14000, 16000, 18000],
    60: [14000, 16000, 18000, 20000],
    70: [15000, 17000, 19000, 21000],
    80: [16500, 18500, 20500, 23000],
    90: [18500, 20500, 23500, 26000],
    100: [22000, 24000, 26000, 29000]
}

plumbing_electrical_thresholds = {
    0: [7000, 8000, 9000, 10000],
    10: [7500, 8500, 9500, 11000],
    20: [8000, 9000, 10000, 11000],
    30: [9000, 10000, 11000, 12000],
    40: [10000, 11000, 12000, 13000],
    50: [11500, 12500, 14000, 15000],
    60: [13000, 14000, 15500, 17000],
    70: [14500, 15500, 17000, 18000],
    80: [15500, 16500, 18000, 19000],
    90: [17000, 18500, 19500, 21000],
    100: [18000, 20000, 22000, 24000]
}

def setup_logging():
    logger = logging.getLogger('commission_calculator')
    logger.setLevel(logging.DEBUG)
    
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # File handler - keeps detailed DEBUG logs
    fh = logging.FileHandler(f'commission_calculations_{current_time}.log')
    fh.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    fh.setFormatter(file_formatter)
    
    # Console handler - only shows critical info and progress
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(message)s')  # Simplified console output
    ch.setFormatter(console_formatter)
    
    # Add filters to prevent duplicate messages
    class ConsoleFilter(logging.Filter):
        def filter(self, record):
            return record.levelno == logging.INFO and not any(x in record.msg for x in [
                'Department Summary',
                'Completed:',
                'Sales:',
                'Total:',
                'Spiffs:',
                'Commissionable Revenue:',
                'Commission:',
                'Processing technician'
            ])
    
    ch.addFilter(ConsoleFilter())
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    
    return logger

def format_currency(amount):
    """Format number as currency string."""
    try:
        if isinstance(amount, str):
            # Remove '$' and ',' from string before converting
            amount = float(amount.replace('$', '').replace(',', ''))
        return f"${amount:,.2f}"
    except (ValueError, TypeError):
        return "$0.00"

def get_user_date() -> datetime:
    while True:
        try:
            date_str = input("Enter a date (mm/dd/yy): ")
            return datetime.strptime(date_str, '%m/%d/%y')
        except ValueError:
            print("Invalid date format. Please use mm/dd/yy format (e.g., 03/15/24)")

def extract_department_number(memo: str) -> int:
    try:
        parts = memo.split('-')[0].strip()
        return int(''.join(filter(str.isdigit, parts)))
    except Exception as e:
        logger.warning(f"Could not extract department number from memo: {memo}")
        return 0

def get_department_with_code(dept_num: int) -> str:
    """
    Returns department name with department code range.
    
    Args:
        dept_num (int): Department number from business unit
        
    Returns:
        str: Formatted string with department name and code range
    """
    if 20 <= dept_num <= 29:
        return 'HVAC (20-29)'
    elif 30 <= dept_num <= 39:
        return 'Plumbing (30-39)'
    elif 40 <= dept_num <= 49:
        return 'Electric (40-49)'
    return 'Unknown (0)'

def get_department_from_number(dept_num: int) -> str:
    if 20 <= dept_num <= 29:
        return 'HVAC'
    elif 30 <= dept_num <= 39:
        return 'Plumbing'
    elif 40 <= dept_num <= 49:
        return 'Electric'
    return 'Unknown'

def autofit_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def extract_department_range(business_unit: str) -> range:
    """Extract department range from business unit string."""
    try:
        unit_number = int(''.join(filter(str.isdigit, business_unit)))
        base = (unit_number // 10) * 10
        if 20 <= base <= 20:  # HVAC
            return range(20, 30)
        elif 30 <= base <= 30:  # Plumbing
            return range(30, 40)
        elif 40 <= base <= 40:  # Electric
            return range(40, 50)
        return range(0, 0)  # Invalid range
    except (ValueError, TypeError):
        logger.warning(f"Could not extract department range from business unit: {business_unit}")
        return range(0, 0)

def is_same_department(source_unit: str, target_unit: str) -> bool:
    """Check if two business units are in the same department range."""
    source_range = extract_department_range(source_unit)
    target_range = extract_department_range(target_unit)
    
    return (len(source_range) > 0 and 
            len(target_range) > 0 and 
            source_range.start == target_range.start)

def format_threshold_scale(thresholds):
    """Format threshold scale as a string showing all tiers."""
    return (f"2%: ${thresholds[0]:,.0f} | "
            f"3%: ${thresholds[1]:,.0f} | "
            f"4%: ${thresholds[2]:,.0f} | "
            f"5%: ${thresholds[3]:,.0f}")

def get_valid_tgls(file_path: str, tech_name: str) -> List[dict]:
    """Process TGL data from Sheet1_TGL to find valid TGLs for a technician."""
    try:
        tgl_df = pd.read_excel(file_path, sheet_name='Sheet1_TGL')
        logger.debug(f"Processing TGLs for {tech_name} from Sheet1_TGL")
        
        valid_tgls = []
        tech_tgls = tgl_df[
            (tgl_df['Lead Generated By'] == tech_name) & 
            (tgl_df['Status'] != 'Canceled')
        ]
        
        for _, tgl in tech_tgls.iterrows():
            source_unit = str(tgl['Business Unit'])
            target_unit = str(tgl['Lead Generated from Business Unit'])
            
            if is_same_department(source_unit, target_unit):
                valid_tgls.append({
                    'job_number': tgl.get('Job #', 'N/A'),
                    'status': tgl['Status'],
                    'business_unit': source_unit,
                    'target_unit': target_unit,
                    'created_date': tgl['Created Date']
                })
                logger.debug(f"Valid TGL found - Job #: {tgl.get('Job #', 'N/A')}, "
                           f"From: {source_unit} To: {target_unit}")
        
        return valid_tgls
        
    except Exception as e:
        logger.error(f"Error processing TGL data for {tech_name}: {str(e)}")
        return []

def get_spiffs_total(file_path: str, tech_name: str) -> tuple[float, dict[str, float]]:
    """
    Get total spiffs amount and department-specific spiffs from Direct Payroll Adjustments.
    Returns tuple of (total_spiffs, department_spiffs_dict)
    """
    try:
        spiffs_df = pd.read_excel(file_path, sheet_name='Direct Payroll Adjustments')
        tech_spiffs = spiffs_df[spiffs_df['Technician'] == tech_name]
        
        spiffs_total = 0
        department_spiffs = {
            'HVAC': 0,
            'Plumbing': 0,
            'Electric': 0
        }
        
        for idx, spiff in tech_spiffs.iterrows():
            amount = float(str(spiff['Amount']).replace('$', '').replace(',', '').strip()) if pd.notnull(spiff['Amount']) else 0
            memo = str(spiff['Memo']).strip() if pd.notnull(spiff['Memo']) else ''
            
            # Skip entries with specific conditions
            if ('tgl' in memo.lower() or 
                '2% commission' in memo.lower() or 
                'reversal of commission' in memo.lower() or
                'commission deduction' in memo.lower()):
                logger.debug(f"Skipping entry: ${amount:,.2f} - Memo: {memo}")
                continue
            
            # Validate memo format
            if not memo[:2].isdigit():
                logger.error(f"Invalid memo format - must start with department number. Row {idx + 2}: {memo}")
                raise ValueError(f"Invalid memo format - must start with department number. Row {idx + 2}: {memo}")
            
            dept_num = int(memo[:2])
            if not (20 <= dept_num <= 29 or 30 <= dept_num <= 39 or 40 <= dept_num <= 49):
                logger.error(f"Invalid department number in memo (must be 20-29, 30-39, or 40-49). Row {idx + 2}: {memo}")
                raise ValueError(f"Invalid department number in memo (must be 20-29, 30-39, or 40-49). Row {idx + 2}: {memo}")
            
            # Check for content after department number
            memo_content = memo[2:].strip().lstrip('-').strip()
            if not memo_content:
                logger.warning(f"Warning: Memo contains no description after department number. Row {idx + 2}: {memo}")
            
            # Add to appropriate department total based on department number
            if 20 <= dept_num <= 29:
                department_spiffs['HVAC'] += amount
                logger.debug(f"Added HVAC spiff: ${amount:,.2f} - Memo: {memo}")
            elif 30 <= dept_num <= 39:
                department_spiffs['Plumbing'] += amount
                logger.debug(f"Added Plumbing spiff: ${amount:,.2f} - Memo: {memo}")
            elif 40 <= dept_num <= 49:
                department_spiffs['Electric'] += amount
                logger.debug(f"Added Electric spiff: ${amount:,.2f} - Memo: {memo}")
            
            spiffs_total += amount
        
        return spiffs_total, department_spiffs
        
    except Exception as e:
        logger.error(f"Error processing spiffs data for {tech_name}: {str(e)}")
        raise  # Re-raise the exception to stop program execution


def calculate_box_metrics(data: pd.DataFrame, tech_name: str) -> Tuple[float, float, float]:
    # Calculate CJR (Box A)
    primary_jobs = data[data['Primary Technician'] == tech_name]
    box_a = primary_jobs['Jobs Total Revenue'].fillna(0).sum()
    
    logger.debug(f"{tech_name} - CJR (Box A) Included Jobs:")
    for _, row in primary_jobs.iterrows():
        logger.debug(f"Invoice: {row.get('Invoice #', 'N/A')}, Revenue: {row['Jobs Total Revenue']}")

    # Calculate TSIS (Box B)
    sold_jobs = data[
        (data['Sold By'] == tech_name) & 
        (data['Primary Technician'] != tech_name)
    ]
    box_b = sold_jobs['Jobs Total Revenue'].fillna(0).sum()
    
    logger.debug(f"{tech_name} - TSIS (Box B) Jobs:")
    for _, row in sold_jobs.iterrows():
        logger.debug(f"Invoice: {row.get('Invoice #', 'N/A')}, Revenue: {row['Jobs Total Revenue']}")

    # Total Revenue (Box C)
    box_c = max(0, box_a + box_b)
    
    logger.debug(f"{tech_name} - Final totals: Box A: ${box_a:,.2f}, Box B: ${box_b:,.2f}, Box C: ${box_c:,.2f}")
    return box_a, box_b, box_c

def calculate_percentages(box_a: float, box_c: float) -> Tuple[int, int]:
    if box_c == 0:
        return 0, 0
        
    raw_scp = (box_a / box_c * 100)
    raw_icp = 100 - raw_scp
    
    raw_scp = max(0, raw_scp)
    raw_icp = max(0, raw_icp)
    
    total = raw_scp + raw_icp
    if total != 0:
        raw_scp = (raw_scp / total) * 100
        raw_icp = (raw_icp / total) * 100
        
    scp = round(raw_scp / 10) * 10
    icp = round(raw_icp / 10) * 10
    
    if scp + icp != 100:
        if scp > icp:
            scp = 100
            icp = 0
        else:
            scp = 0
            icp = 100
    
    return int(scp), int(icp)

def calculate_average_ticket_value(data: pd.DataFrame, tech_name: str) -> float:
    qualifying_jobs = data[
        (data['Primary Technician'] == tech_name)
    ]
    
    logger.debug(f"Average ticket calculation for {tech_name}:")
    logger.debug(f"Total jobs as Primary Tech: {len(data[data['Primary Technician'] == tech_name])}")
    
    if len(qualifying_jobs) == 0:
        logger.debug(f"No qualifying True Opportunity jobs found for {tech_name}")
        return 0
    
    total_revenue = qualifying_jobs['Jobs Total Revenue'].fillna(0).sum()
    avg_ticket = total_revenue / len(qualifying_jobs)
    
    logger.debug(f"Total revenue from True Opportunity jobs: ${total_revenue:,.2f}")
    logger.debug(f"Average ticket: ${round(avg_ticket, 2):,.2f}")
    
    return round(avg_ticket, 2)

def get_excused_hours(file_path: str, sheet_name: str = '2024') -> Dict[str, int]:
    try:
        base_date = get_user_date()
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        start_of_week = base_date - timedelta(days=base_date.weekday())
        end_of_week = start_of_week + timedelta(days=4)
        
        def format_date(date):
            day = date.day
            if 4 <= day <= 20 or 24 <= day <= 30:
                suffix = "th"
            else:
                suffix = ["st", "nd", "rd"][day % 10 - 1]
            return f"{date.strftime('%B')} {day}{suffix}"
        
        target_week = f"{format_date(start_of_week)} - {format_date(end_of_week)}"
        logger.debug(f"Looking for week range: {target_week}")
        
        start_col = None
        for row_idx in [0, 1]:
            for col_idx, header in enumerate(df.iloc[row_idx]):
                if isinstance(header, str) and target_week.strip() == header.strip():
                    start_col = col_idx
                    logger.debug(f"Found target week in row {row_idx}, column {col_idx}")
                    break
            if start_col is not None:
                break
        
        if start_col is None:
            logger.warning(f"Week range '{target_week}' not found in time off sheet")
            return {}
        
        week_columns = list(range(start_col, start_col + 5))
        hours_summary = {}
        
        for index in range(2, len(df)):
            row = df.iloc[index]
            technician_name = row[0]
            
            if pd.isna(technician_name) or str(technician_name).strip() == '':
                continue
            
            technician_name = str(technician_name).strip()
            total_hours = 0
            
            for col in week_columns:
                try:
                    cell_value = str(row[col]).strip().lower()
                    if cell_value in ['x', 'r', 'v']:
                        total_hours += 8
                        logger.debug(f"Found time off marker '{cell_value}' for {technician_name} in column {col}")
                except Exception as e:
                    logger.warning(f"Error processing cell for {technician_name} in column {col}: {str(e)}")
                    continue
            
            if total_hours > 0:
                hours_summary[technician_name] = total_hours
                logger.debug(f"Found {total_hours} excused hours for {technician_name}")
        
        return hours_summary
        
    except Exception as e:
        logger.error(f"Error analyzing time off data: {str(e)}")
        return {}

def calculate_department_revenue(data: pd.DataFrame, tech_name: str) -> Dict[str, Dict[str, float]]:
    """
    Calculate revenue by department for both completed jobs and sales.
    Also calculate the combined total per department for commission purposes.
    """
    revenue_by_dept = {
        'completed': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0},
        'sales': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0},
        'combined': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0}
    }
    
    # Track completed jobs (where they are primary tech)
    completed_jobs = data[data['Primary Technician'] == tech_name]
    for _, job in completed_jobs.iterrows():
        dept_num = extract_department_number(str(job.get('Business Unit', '')))
        dept = get_department_from_number(dept_num)
        revenue = job.get('Jobs Total Revenue', 0) or 0
        revenue_by_dept['completed'][dept] += revenue
        revenue_by_dept['combined'][dept] += revenue
        logger.debug(f"{tech_name} - Completed job in {dept}: ${revenue:,.2f}")

    # Track sales (where they sold but didn't complete)
    sold_jobs = data[
        (data['Sold By'] == tech_name) & 
        (data['Primary Technician'] != tech_name)
    ]
    for _, job in sold_jobs.iterrows():
        dept_num = extract_department_number(str(job.get('Business Unit', '')))
        dept = get_department_from_number(dept_num)
        revenue = job.get('Jobs Total Revenue', 0) or 0
        revenue_by_dept['sales'][dept] += revenue
        revenue_by_dept['combined'][dept] += revenue
        logger.debug(f"{tech_name} - Sold job in {dept}: ${revenue:,.2f}")

    # Log summary for the technician
    for category in ['completed', 'sales', 'combined']:
        logger.debug(f"{tech_name} - {category.title()} revenue by department:")
        for dept, amount in revenue_by_dept[category].items():
            if amount > 0:
                logger.debug(f"  {dept}: ${amount:,.2f}")

    return revenue_by_dept

def get_commission_rate(total_revenue: float, flipped_percent: float, department: str, 
                       excused_hours: int, tgl_reduction: float):
    logger.debug(f"\nDetailed threshold calculation:")
    
    flipped_percent = min(100, max(0, int(round(flipped_percent / 10) * 10)))
    
    if department in ['Electric', 'Plumbing']:
        thresholds = plumbing_electrical_thresholds
    else:
        thresholds = hvac_thresholds
    
    tier_thresholds = thresholds[flipped_percent].copy()
    logger.debug(f"Using {department} thresholds for ICP {flipped_percent}%: {tier_thresholds}")
    
    days_off = min(5, excused_hours / 8)
    reduction_factor = max(0, 1 - (0.20 * days_off))
    logger.debug(f"Time off reduction factor: {reduction_factor} ({days_off} days)")
    
    # First apply time off reduction
    time_off_adjusted = [threshold * reduction_factor for threshold in tier_thresholds]
    logger.debug(f"After time off adjustment: {time_off_adjusted}")
    
    # Then apply TGL reduction
    adjusted_thresholds = [max(0, threshold - tgl_reduction) for threshold in time_off_adjusted]
    logger.debug(f"After TGL reduction: {adjusted_thresholds}")
    
    # Determine commission rate based on highest threshold met
    if total_revenue >= adjusted_thresholds[3]:
        rate = 0.05
    elif total_revenue >= adjusted_thresholds[2]:
        rate = 0.04
    elif total_revenue >= adjusted_thresholds[1]:
        rate = 0.03
    elif total_revenue >= adjusted_thresholds[0]:
        rate = 0.02
    else:
        rate = 0
    
    logger.debug(f"Final commission rate: {rate}")
    return rate, adjusted_thresholds, tier_thresholds

def format_department_revenue(revenue_data: Dict[str, Dict[str, float]], 
                            commission_rate: float,
                            department_spiffs: Dict[str, float]) -> Dict[str, str]:
    """
    Format department revenue data into strings for Excel output, including commission calculations.
    Takes into account department-specific spiffs when calculating commission.
    """
    formatted = {}
    
    for dept in ['HVAC', 'Plumbing', 'Electric']:
        # Format revenue and sales
        completed = revenue_data['completed'][dept]
        sales = revenue_data['sales'][dept]
        combined = revenue_data['combined'][dept]
        
        # Deduct department-specific spiffs from combined revenue
        dept_spiffs = department_spiffs[dept]
        adjusted_combined = combined - dept_spiffs
        
        formatted[f"{dept} Completed"] = f"${completed:,.2f}"
        formatted[f"{dept} Sales"] = f"${sales:,.2f}"
        formatted[f"{dept} Total"] = f"${adjusted_combined:,.2f}"  # Use adjusted total
        formatted[f"{dept} Commission"] = f"${(adjusted_combined * commission_rate):,.2f}"  # Use calculated commission rate
        
        logger.debug(f"{dept} breakdown:")
        logger.debug(f"  Combined Revenue: ${combined:,.2f}")
        logger.debug(f"  Department Spiffs: ${dept_spiffs:,.2f}")
        logger.debug(f"  Adjusted Total: ${adjusted_combined:,.2f}")
        logger.debug(f"  Commission Rate: {commission_rate*100}%")
        logger.debug(f"  Final Commission: ${(adjusted_combined * commission_rate):,.2f}")
    
    return formatted

def main():
    
    technicians_to_track = [
        "Andrew Wycoff", "Andy Ventura", "Artie Straniti", "Brett Allen", "Carter Bruce",
        "Chris Smith", "David Franklin", "David Knox",
        "Ethan Ficklin", "Glenn Griffin", "Hunter Stanley", 
        "Jacob Simpson", "Jake West", "Jason Kerns", "Josue Rodriguez", "Justin Barron", 
        "Kevin Stanley", "Pablo Silvas", "Patrick Bowerman", "Robert McGhee", "Ronnie Bland", 
        "Sean Lynders", "Shawn Hollingsworth", "Stephen Starner", "Thomas Shawaryn", 
        "Tim Kulesza", "WT Settle", "Will Winfree"
    ]

    file_path = r'C:\Users\abatlouni\Downloads\combined_data.xlsx'
    logger.info(f"Loading data from {file_path}")
    
    try:
        data = pd.read_excel(file_path, sheet_name='Sheet1')
        tech_data = pd.read_excel(file_path, sheet_name='Sheet1_Tech')
    except Exception as e:
        logger.error(f"Error loading Excel sheets: {e}")
        sys.exit(1)
    
    logger.debug(f"Columns in 'Sheet1_Tech': {tech_data.columns.tolist()}")
    
    excused_hours_dict = get_excused_hours(file_path)

    required_columns = ['Name', 'Technician Business Unit']
    missing_columns = [col for col in required_columns if col not in tech_data.columns]
    
    if missing_columns:
        logger.error(f"Missing columns in 'Sheet1_Tech': {missing_columns}")
        sys.exit(1)
    
    tech_dept_map = dict(zip(tech_data['Name'], tech_data['Technician Business Unit']))

    results = []

    for tech_name in technicians_to_track:
        logger.info(f"\nProcessing technician: {tech_name}")
        
        # Basic calculations
        box_a, box_b, box_c = calculate_box_metrics(data, tech_name)
        scp, icp = calculate_percentages(box_a, box_c)
        
        # Calculate department-specific revenue
        dept_revenue = calculate_department_revenue(data, tech_name)
        
        # Get spiffs with department breakdown
        spiffs_total, department_spiffs = get_spiffs_total(file_path, tech_name)
        
        # Get valid TGLs
        valid_tgls = get_valid_tgls(file_path, tech_name)
        
        # Calculate average ticket
        avg_ticket = calculate_average_ticket_value(data, tech_name)
        
        # Get department
        dept_unit_str = tech_dept_map.get(tech_name, '0')
        dept_num = extract_department_number(dept_unit_str) if isinstance(dept_unit_str, str) else 0
        department = get_department_from_number(dept_num)  # Keep this for internal logic
        department_with_code = get_department_with_code(dept_num)  # New formatted version

        # Calculate TGL threshold reduction
        tgl_reduction = avg_ticket * len(valid_tgls) if avg_ticket > 0 else 0
        
        excused_hours = excused_hours_dict.get(tech_name, 0)
        
        # Calculate commission rate using ICP
        commission_rate, adjusted_thresholds, base_thresholds = get_commission_rate(
            box_c, icp, department, excused_hours, tgl_reduction
        )
        
        # Format threshold scales
        base_threshold_scale = format_threshold_scale(base_thresholds)
        adjusted_threshold_scale = format_threshold_scale(adjusted_thresholds)
        
        # Calculate department-specific commissions with spiff reductions
        formatted_dept_data = format_department_revenue(
            dept_revenue,
            commission_rate,
            department_spiffs
        )
        
        # Calculate total commissionable revenue (excluding spiffs)
        commissionable_revenue = box_c - spiffs_total
        final_commission = commissionable_revenue * commission_rate
        
        # Build result dictionary with new department revenue and commission data
        result = {
            # Identifying Info
            'Technician': tech_name,
            'Main Dept': department_with_code,
            
            # Revenue Components
            'Total Revenue': box_c,
            'Completed Job Revenue': box_a,
            'Tech-Sourced Install Sales': box_b,
            
            # Department-specific Revenue and Commission
            'HVAC Revenue': formatted_dept_data['HVAC Completed'],
            'HVAC Sales': formatted_dept_data['HVAC Sales'],
            'HVAC Spiffs': format_currency(department_spiffs['HVAC']),
            'HVAC Total': formatted_dept_data['HVAC Total'],
            'HVAC Commission': formatted_dept_data['HVAC Commission'],

            
            'Plumbing Revenue': formatted_dept_data['Plumbing Completed'],
            'Plumbing Sales': formatted_dept_data['Plumbing Sales'],
            'Plumbing Spiffs': format_currency(department_spiffs['Plumbing']),
            'Plumbing Total': formatted_dept_data['Plumbing Total'],
            'Plumbing Commission': formatted_dept_data['Plumbing Commission'],
            
            'Electric Revenue': formatted_dept_data['Electric Completed'],
            'Electric Sales': formatted_dept_data['Electric Sales'],
            'Electric Spiffs': format_currency(department_spiffs['Electric']),
            'Electric Total': formatted_dept_data['Electric Total'],
            'Electric Commission': formatted_dept_data['Electric Commission'],
            
            # Performance Percentages
            'Service Completion %': scp,
            'Install Contribution %': icp,
            
            # Adjustments & Qualifiers
            'Excused Hours': excused_hours,
            'Spiffs': spiffs_total,
            'Valid TGLs': len(valid_tgls),
            'Avg Ticket $': round(avg_ticket, 2),
            'TGL Threshold Reduction': tgl_reduction,
            
            # Commission Calculation
            'Base Threshold Scale': base_threshold_scale,
            'Adjusted Threshold Scale': adjusted_threshold_scale,
            'Commissionable Revenue': commissionable_revenue,
            'Commission Rate %': commission_rate * 100,
            'Total Commission': round(final_commission, 2),
            
            # Qualification Status
            'Status': f"Qualified for {commission_rate*100}% tier" if commission_rate > 0 else "Did not qualify"
        }
        
        results.append(result)

    # Create DataFrame and write to Excel
    results_df = pd.DataFrame(results)
    output_path = r'C:\Users\abatlouni\Downloads\output.xlsx'
    logger.info(f"\nWriting results to {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Technician Revenue Totals', index=False)
        wb = writer.book
        ws = wb['Technician Revenue Totals']
        autofit_columns(ws)

    logger.info("Commission calculations completed")

if __name__ == "__main__":
    logger = setup_logging()
    main()