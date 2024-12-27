import pandas as pd
import logging
import sys
import os
import glob
import shutil
import re
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, List
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from pathlib import Path
from collections import defaultdict

logger = logging.getLogger('commission_processor')

# Constants
LOCATION_ID = 'L100'
COMPANY_CODE = 'J6P'

# Column order for output
COLUMN_ORDER = [
    'Badge ID', 'Technician', 'Main Dept',
    # HVAC Department Subdepartments
    '20 Revenue', '20 Sales', '20 Total', '20 Spiffs', '20 Commission',
    '21 Revenue', '21 Sales', '21 Total', '21 Spiffs',  '21 Commission',
    '22 Revenue', '22 Sales', '22 Total', '22 Spiffs', '22 Commission',
    '24 Revenue', '24 Sales', '24 Total', '24 Spiffs', '24 Commission',
    '25 Revenue', '25 Sales', '25 Total', '25 Spiffs', '25 Commission',
    '27 Revenue', '27 Sales', '27 Total', '27 Spiffs', '27 Commission',
    # HVAC Department Totals
    'HVAC Revenue', 'HVAC Sales', 'HVAC Spiffs', 'HVAC Total', 'HVAC Commission',
    # Plumbing Department Subdepartments
    '30 Revenue', '30 Sales', '30 Total', '30 Spiffs', '30 Commission',
    '31 Revenue', '31 Sales', '31 Total', '31 Spiffs', '31 Commission',
    '33 Revenue', '33 Sales', '33 Total', '33 Spiffs', '33 Commission',
    '34 Revenue', '34 Sales', '34 Total', '34 Spiffs', '34 Commission',
    # Plumbing Department Totals
    'Plumbing Revenue', 'Plumbing Sales', 'Plumbing Spiffs', 'Plumbing Total', 'Plumbing Commission',
    # Electric Department Subdepartments
    '40 Revenue', '40 Sales', '40 Total', '40 Spiffs', '40 Commission',
    '41 Revenue', '41 Sales', '41 Total', '41 Spiffs', '41 Commission',
    '42 Revenue', '42 Sales', '42 Total', '42 Spiffs', '42 Commission',
    # Electric Department Totals
    'Electric Revenue', 'Electric Sales', 'Electric Spiffs', 'Electric Total', 'Electric Commission',
    # Performance Metrics
    'Completed Job Revenue', 'Tech-Sourced Install Sales',
    'Service Completion %', 'Install Contribution %',
    'Excused Hours', 'Spiffs', 'Valid TGLs', 'Avg Ticket $',
    'TGL Threshold Reduction', 'Base Threshold Scale', 'Adjusted Threshold Scale',
    'Commission Rate %', 'Total Revenue', 'Commissionable Revenue',
    'Status', 'Total Commission'
]

# Define threshold tables
HVAC_THRESHOLDS = {
    0: [7000, 8000, 9000, 10000],
    10: [7500, 8500, 9500, 11000],
    20: [8000, 9000, 10000, 11000],
    30: [9000, 10000, 11000, 12000],
    40: [10000, 11000, 12000, 13000],
    50: [12000, 14000, 16000, 18000],
    60: [14000, 16000, 18000, 20000],
    70: [15000, 17000, 19000, 21000],
    80: [16500, 18500, 20500, 23000],
    90: [18500, 20500, 23500, 26000],
    100: [22000, 24000, 26000, 29000]
}

PLUMBING_ELECTRICAL_THRESHOLDS = {
    0: [7000, 8000, 9000, 10000],
    10: [7500, 8500, 9500, 11000],
    20: [8000, 9000, 10000, 11000],
    30: [9000, 10000, 11000, 12000],
    40: [10000, 11000, 12000, 13000],
    50: [11500, 12500, 14000, 15000],
    60: [13000, 14000, 15500, 17000],
    70: [14500, 15500, 17000, 18000],
    80: [15500, 16500, 18000, 19000],
    90: [17000, 18500, 19500, 21000],
    100: [18000, 20000, 22000, 24000]
}

# Department code mapping
DEPARTMENT_CODES = {
    '20': {'code': '2000000', 'desc': 'HVAC SERVICE'},
    '21': {'code': '2100000', 'desc': 'HVAC INSTALL'},
    '22': {'code': '2200000', 'desc': 'MAINTENANCE MVP'},
    '24': {'code': '2400000', 'desc': 'OIL SERVICE'},
    '25': {'code': '2500000', 'desc': 'OIL MAINTENANCE MVP'},
    '27': {'code': '2700000', 'desc': 'HVAC DUCT CLEANING'},
    '30': {'code': '3000000', 'desc': 'PLUMBING SERVICE'},
    '31': {'code': '3100000', 'desc': 'PLUMBING INSTALL'},
    '33': {'code': '3300000', 'desc': 'PLUMBING DRAIN CLEANING'},
    '34': {'code': '3400000', 'desc': 'PLUMBING EXCAVATION'},
    '40': {'code': '4000000', 'desc': 'ELECTRICAL SERVICE'},
    '41': {'code': '4100000', 'desc': 'ELECTRICAL INSTALL'},
    '42': {'code': '4200000', 'desc': 'GENERATOR MAINTENANCE'}
}


class DateValidator:
    @staticmethod
    def parse_filename_date_range(filename: str) -> Optional[Tuple[datetime, datetime]]:
        """Extract date range from filename patterns like 'MM_DD_YY - MM_DD_YY'"""
        pattern = r'(\d{2}_\d{2}_\d{2})\s*-\s*(\d{2}_\d{2}_\d{2})'
        match = re.search(pattern, filename)
        
        if not match:
            return None
            
        try:
            start_str, end_str = match.groups()
            start_date = datetime.strptime(start_str, '%m_%d_%y')
            end_date = datetime.strptime(end_str, '%m_%d_%y')
            return start_date, end_date
        except ValueError:
            return None

    @staticmethod
    def get_week_range(date: datetime, show_message: bool = False) -> Tuple[datetime, datetime]:
        """
        Get the Monday-Sunday dates for the week containing the input date.
        Any date entered will map to its corresponding week's Monday-Sunday range.
        """
        # Get Monday of the week (subtract days until we hit Monday)
        start_of_week = date - timedelta(days=date.weekday())
        # Get Sunday (add 6 days to Monday)
        end_of_week = start_of_week + timedelta(days=6)
        
        if show_message:
            print(f"\nInput date {date.strftime('%m/%d/%y')} ({date.strftime('%A')})")
            print(f"Maps to week: {start_of_week.strftime('%m/%d/%y')} (Monday) - {end_of_week.strftime('%m/%d/%y')} (Sunday)")
        
        return start_of_week, end_of_week

    @staticmethod
    def format_date_for_comparison(date: datetime) -> str:
        """Format date as MM_DD_YY for string comparison"""
        return date.strftime('%m_%d_%y')

    @staticmethod
    def format_date_for_display(date: datetime) -> str:
        """Format date as MM/DD/YY for display"""
        return date.strftime('%m/%d/%y')

    @classmethod
    def analyze_uuid_file_dates(cls, file_path: str) -> Optional[Tuple[datetime, datetime]]:
        """Analyze UUID file dates from Direct Payroll Adjustments sheet"""
        try:
            df = pd.read_excel(file_path, sheet_name='Direct Payroll Adjustments')
            
            if 'Posted On' not in df.columns:
                return None
                
            if not pd.api.types.is_datetime64_any_dtype(df['Posted On']):
                df['Posted On'] = pd.to_datetime(df['Posted On'])
            
            # Get min and max dates from Posted On column, excluding NaN values
            valid_dates = df['Posted On'].dropna()
            if valid_dates.empty:
                return None
                
            earliest_date = valid_dates.min()
            latest_date = valid_dates.max()
            
            print(f"\n\"Posted On\"column in UUID file date range: {earliest_date.strftime('%m/%d/%y')} to {latest_date.strftime('%m/%d/%y')}")
            
            return earliest_date.to_pydatetime(), latest_date.to_pydatetime()
            
        except Exception as e:
            print(f"\nDEBUG: Error in analyze_uuid_file_dates: {str(e)}")
            return None

    @classmethod
    def validate_files_for_date(cls, directory: str, user_date: datetime) -> Tuple[bool, List[str], Optional[List[str]], dict]:
        start_week, end_week = cls.get_week_range(user_date, show_message=True)
        expected_start = cls.format_date_for_comparison(start_week)
        expected_end = cls.format_date_for_comparison(end_week)

        errors = []
        found_files = {}

        file_patterns = {
            'tech': f"Technician Department_Dated {expected_start} - {expected_end}.xlsx",
            'tgl': f"TGLs Set _Dated {expected_start} - {expected_end}.xlsx",
            'jobs': f"Copy of Jobs Report for Performance -DE2_Dated {expected_start} - {expected_end}.xlsx"
        }

        # Check each expected file
        for file_type, expected_pattern in file_patterns.items():
            matching_files = glob.glob(os.path.join(directory, expected_pattern))
            
            if not matching_files:
                # Check if files exist with different dates
                pattern_base = expected_pattern.split('Dated')[0] + "Dated *"
                existing_files = glob.glob(os.path.join(directory, pattern_base))
                if existing_files:
                    # Found files but wrong week
                    example_file = os.path.basename(existing_files[0])
                    date_range = cls.parse_filename_date_range(example_file)
                    if date_range:
                        actual_start, actual_end = date_range
                        errors.append(
                            f"Found {file_type} file for different week "
                            f"({cls.format_date_for_display(actual_start)} - "
                            f"{cls.format_date_for_display(actual_end)})"
                        )
                    else:
                        errors.append(f"Found {file_type} file but couldn't parse its date range")
                else:
                    errors.append(f"Missing {file_type} file for week of {cls.format_date_for_display(start_week)}")
            else:
                found_files[file_type] = matching_files[0]

        # Check UUID files
        uuid_files = glob.glob(os.path.join(directory, "????????-????-????-????-????????????.xlsx"))
        if not uuid_files:
            errors.append("Missing UUID file")
            return len(errors) == 0, errors, None, {}
        
        # Time off file
        time_off_file = os.path.join(directory, "Approved_Time_Off 2023.xlsx")
        if not os.path.exists(time_off_file):
            errors.append("Missing Time Off file")
        else:
            found_files['time_off'] = time_off_file

        return len(errors) == 0, errors, uuid_files, found_files

    @classmethod
    def validate_files_for_date_with_uuid(cls, directory: str, user_date: datetime, selected_uuid: str) -> Tuple[bool, List[str]]:
        """Validate files with a specific UUID file."""
        is_valid, errors, uuid_files, found_files_dummy = cls.validate_files_for_date(directory, user_date)
        
        # Remove any UUID-related errors as we're using a specific one
        errors = [e for e in errors if not ("UUID file" in e)]
        
        # Validate the selected UUID file
        start_week, end_week = cls.get_week_range(user_date)
        uuid_dates = cls.analyze_uuid_file_dates(selected_uuid)
        
        if uuid_dates:
            uuid_start, uuid_end = uuid_dates

            # Use the same mid-week coverage requirement as in validate_files_for_date
            mid_week_start = start_week + timedelta(days=2)  # Wednesday
            mid_week_end = start_week + timedelta(days=4)    # Friday

            if not (uuid_start.date() <= mid_week_start.date() and uuid_end.date() >= mid_week_end.date()):
                errors.append(
                    f"Selected UUID file data range ({cls.format_date_for_display(uuid_start)} - "
                    f"{cls.format_date_for_display(uuid_end)}) "
                    f"does not cover the mid-week (Wed-Fri)."
                )
        
        return len(errors) == 0, errors

def get_validated_user_date(base_path: str) -> Tuple[datetime, str, dict]:
    validator = DateValidator()

    def get_uuid_file_choice(uuid_files: List[str]) -> str:
        most_recent = os.path.basename(max(uuid_files, key=os.path.getmtime))
        print(f"\n- Multiple UUID files found. Will use most recent: {most_recent}")
        
        while True:
            choice = input("\nIs this acceptable? (Y/N): ").strip().upper()
            if choice in ['Y', 'N']:
                break
            print("Invalid input. Please enter Y or N.")
        
        if choice == 'Y':
            return max(uuid_files, key=os.path.getmtime)
        
        while True:
            print("\nPlease either:")
            print("1. Enter the name of the UUID file from within your Downloads folder")
            print("2. Type \"exit\", or press ctrl+c to exit program")
            filename = input().strip()
                            
            if filename.lower() == 'exit':
                print("\nExiting program. Please ensure you have the correct files and try again.")
                sys.exit(0)

            matching_files = [f for f in uuid_files if os.path.basename(f) == filename]
            if not matching_files:
                print(f"\nError: {filename} not found in Downloads folder.")
                print("Available UUID files:")
                for f in uuid_files:
                    print(f"- {os.path.basename(f)}")
                continue
            
            confirm = input(f"\nYou entered {filename}, is this correct? (Y/N): ").strip().upper()
            if confirm == 'Y':
                return matching_files[0]
            elif confirm == 'N':
                continue
            else:
                print("Invalid input. Please enter Y or N.")

    while True:
        print("\nEnter a date (mm/dd/yy) or 'exit' to quit: ")
        user_input = input().strip()
        
        if user_input.lower() == 'exit':
            print("\nExiting program. Please ensure you have the correct files and try again.")
            sys.exit(0)
        
        try:
            date = datetime.strptime(user_input, '%m/%d/%y')
        except ValueError:
            print("Invalid date format. Please use mm/dd/yy format (e.g., 12/22/24)")
            continue
        
        is_valid, errors, uuid_files, found_files = validator.validate_files_for_date(base_path, date)
        
        if uuid_files and len(uuid_files) > 1:
            selected_uuid_file = get_uuid_file_choice(uuid_files)
            found_files['uuid'] = selected_uuid_file
            is_valid, errors = validator.validate_files_for_date_with_uuid(base_path, date, selected_uuid_file)
        elif uuid_files and len(uuid_files) == 1:
            found_files['uuid'] = uuid_files[0]

        if is_valid and 'uuid' in found_files:
            print("All required files found with correct date ranges!")
            return date, found_files['uuid'], found_files
        else:
            print("\nFile validation errors found:")
            for error in errors:
                print(f"- {error}")
            print("\nPlease either:")
            print("1. Enter a different date")
            print("2. Download the correct files for this date")
            print("3. Type 'exit' to quit the program")
            continue


SUBDEPARTMENT_MAP = {
    '00': '00 - ADMINISTRATIVE',
    '41': '41 - ELECTRICAL INSTALL',
    '40': '40 - ELECTRICAL SERVICE',
    '42': '42 - GENERATOR MAINTENANCE',
    '27': '27 - HVAC DUCT CLEANING',
    '21': '21 - HVAC INSTALL',
    '23': '23 - HVAC SALES',
    '20': '20 - HVAC SERVICE',
    '22': '22 - MAINTENANCE MVP',
    '25': '25 - OIL MAINTENANCE MVP',
    '24': '24 - OIL SERVICE',
    '33': '33 - PLUMBING DRAIN CLEANING',
    '34': '34 - PLUMBING EXCAVATION',
    '31': '31 - PLUMBING INSTALL',
    '32': '32 - PLUMBING MAINTENANCE',
    '30': '30 - PLUMBING SERVICE'
}

EXCLUDED_TECHS = [
    "Michael Appleton",
    "Bill Dooly",
    "David Forney", 
    "Dave Elphee",
    "Jim Pumphrey",
    "David Franklin",
    "Mike Wright",
    "Devyn Hitt",
    "Stuart Deary",
    "Jason Knight",
    "Larry Armell",
    "Chris Smith",
    "Will Winfree",
    "Trey Holt III",
    "Gilberto Corvetto"

]

def format_badge_id(badge_id):
    """Clean and return the payroll ID with THREE leading zeros."""
    if pd.isna(badge_id):
        return None
    # Convert to string, remove any decimal points, and ensure THREE leading zeros
    badge_str = str(badge_id).split('.')[0]
    return badge_str.zfill(9)  # Changed from 8 to 9 to ensure THREE leading zeros

def determine_tech_type(business_unit: str) -> str:
    """
    Determine technician type based on business unit.
    
    Returns:
        str: One of three values:
        - 'ADMIN' for administrative/sales staff (to be disregarded)
        - 'SERVICE' for service technicians (PCM, eligible for paystats and spiffs/TGL)
        - 'INSTALL' for installers (ICM, eligible for GP and spiffs/TGL)
    """
    try:
        if pd.isna(business_unit):
            logger.debug(f"Empty business unit, defaulting to ADMIN")
            return 'ADMIN'
            
        bu_upper = str(business_unit).upper()
        logger.debug(f"Checking business unit: {bu_upper}")
        
        # First check for administrative/sales units
        admin_patterns = ['ADMINISTRATIVE', '23 -', 'SALES']
        if any(pattern in bu_upper for pattern in admin_patterns):
            logger.debug(f"Business unit {business_unit} identified as ADMIN")
            return 'ADMIN'
            
        # Then check for service units
        service_patterns = ['SERVICE']
        if any(pattern in bu_upper for pattern in service_patterns):
            logger.debug(f"Business unit {business_unit} identified as SERVICE")
            return 'SERVICE'
            
        # Everything else (INSTALL, EXCAVATION, etc.) is considered an installer
        logger.debug(f"Business unit {business_unit} identified as INSTALL")
        return 'INSTALL'
        
    except Exception as e:
        logger.error(f"Error determining tech type for business unit {business_unit}: {str(e)}")
        return 'ADMIN'  # Default to ADMIN in case of errors

# Dataclass definitions
@dataclass
class PayrollEntry:
    company_code: str = COMPANY_CODE
    badge_id: str = ''
    date: str = ''
    amount: float = 0.0
    pay_code: str = ''  # Valid codes: PCM (Service Tech), ICM (Installer), SPF (Spiffs/TGL)
    dept: str = ''
    location_id: str = LOCATION_ID

    def __post_init__(self):
        """Validate pay code after initialization."""
        valid_codes = {'PCM', 'ICM', 'SPF'}
        if self.pay_code and self.pay_code not in valid_codes:
            raise ValueError(f"Invalid pay code: {self.pay_code}. Must be one of: {', '.join(valid_codes)}")

def get_main_department_code(subdept_code: str) -> str:
    """
    Get the main department code (20/30/40) from a subdepartment code.
    
    Args:
        subdept_code (str): Two-digit subdepartment code (e.g., '21', '31', '42')
        
    Returns:
        str: Main department code ('20', '30', or '40')
    """
    try:
        first_digit = str(subdept_code)[0]
        if first_digit == '2':
            return '20'
        elif first_digit == '3':
            return '30'
        elif first_digit == '4':
            return '40'
        return '00'
    except (IndexError, TypeError):
        return '00'

def get_full_department_code(subdept_code: str) -> str:
    """
    Get the full 7-digit department code based on main department.
    
    Args:
        subdept_code (str): Two-digit subdepartment code
        
    Returns:
        str: Seven-digit department code (e.g., '2000000', '3000000', '4000000')
    """
    main_dept = get_main_department_code(subdept_code)
    return DEPARTMENT_CODES.get(main_dept, {'code': '0000000'})['code']

def get_tech_home_department(tech_business_unit: str) -> str:
    """
    Get a technician's home department code from their business unit.
    E.g., "PLUMBING SERVICE 30" -> "3000000"
    """
    try:
        # Extract numbers from the business unit
        numbers = ''.join(filter(str.isdigit, str(tech_business_unit)))
        if len(numbers) >= 2:
            first_two = numbers[:2]
            first_digit = first_two[0]
            
            # Map to 7-digit department code
            if first_digit == '2':
                return '2000000'  # HVAC
            elif first_digit == '3':
                return '3000000'  # Plumbing
            elif first_digit == '4':
                return '4000000'  # Electric
        return '0000000'
    except (IndexError, AttributeError):
        return '0000000'

def get_service_department_code(subdept_code: str) -> str:
    """
    Get the full 7-digit service department code based on main department.
    Used for PCM entries and TGLs.
    
    Args:
        subdept_code (str): Two-digit subdepartment code
        
    Returns:
        str: Seven-digit department code (e.g., 2000000, 3000000, 4000000)
    """
    main_dept = get_main_department_code(subdept_code)
    return DEPARTMENT_CODES.get(main_dept, {'code': '0000000'})['code']

def consolidate_negative_spiffs(tech_data: pd.DataFrame, spiffs_df: pd.DataFrame) -> Tuple[List[dict], List[dict]]:
    """
    Consolidate negative spiffs by technician's home department.
    
    Args:
        tech_data (pd.DataFrame): DataFrame containing technician information
        spiffs_df (pd.DataFrame): DataFrame containing spiff entries
        
    Returns:
        Tuple[List[dict], List[dict]]: Tuple containing:
            - List of consolidated negative spiff entries
            - List of remaining positive spiff entries
    """
    negative_entries = []
    positive_entries = []
    
    # Group spiffs by technician
    for tech_name, tech_spiffs in spiffs_df.groupby('Technician'):
        try:
            # Get technician's home department
            tech_info = tech_data[tech_data['Name'] == tech_name]
            if tech_info.empty:
                continue
                
            home_dept = get_tech_home_department(tech_info.iloc[0]['Technician Business Unit'])
            badge_id = tech_info.iloc[0]['Badge ID']
            
            # Calculate total negative spiffs
            total_negative = 0
            for _, spiff in tech_spiffs.iterrows():
                try:
                    amount = float(str(spiff['Amount']).replace('$', '').replace(',', ''))
                    if amount < 0:
                        total_negative += amount
                except (ValueError, TypeError):
                    continue
            
            if total_negative < 0:
                negative_entries.append({
                    'Technician': tech_name,
                    'Badge ID': badge_id,
                    'Home Department': home_dept,
                    'Amount': total_negative,
                    'Type': 'Consolidated Negative'
                })
            
            # Keep positive spiffs separate
            for _, spiff in tech_spiffs.iterrows():
                try:
                    amount = float(str(spiff['Amount']).replace('$', '').replace(',', ''))
                    if amount > 0:
                        positive_entries.append({
                            'Technician': tech_name,
                            'Badge ID': badge_id,
                            'Department': get_service_department_code(extract_dept_code(spiff['Memo'])),
                            'Amount': amount,
                            'Memo': spiff['Memo'],
                            'Type': 'Positive'
                        })
                except (ValueError, TypeError):
                    continue
                    
        except Exception as e:
            logger.error(f"Error processing spiffs for {tech_name}: {str(e)}")
            continue
            
    return negative_entries, positive_entries

def process_pcm_entry(tech_data: pd.DataFrame, tech_name: str, subdept_code: str, 
                     amount: float, date: str) -> Optional[PayrollEntry]:
    """
    Create a PCM entry using the service department's main code.
    
    Args:
        tech_data (pd.DataFrame): DataFrame containing technician information
        tech_name (str): Name of the technician
        subdept_code (str): Two-digit subdepartment code where service was performed
        amount (float): Commission amount
        date (str): Target date for the entry
        
    Returns:
        Optional[PayrollEntry]: PayrollEntry object or None if invalid
    """
    try:
        # Get technician info
        tech_info = tech_data[tech_data['Name'] == tech_name]
        if tech_info.empty:
            return None
            
        badge_id = tech_info.iloc[0]['Badge ID']
        if pd.isna(badge_id):
            return None
            
        # Use the service department's main code
        dept_code = get_full_department_code(subdept_code)
        
        return PayrollEntry(
            company_code=COMPANY_CODE,
            badge_id=badge_id,
            date=date,
            amount=amount,
            pay_code='PCM',
            dept=dept_code,
            location_id=LOCATION_ID
        )
        
    except Exception as e:
        logger.error(f"Error creating PCM entry for {tech_name}: {str(e)}")
        return None

# Utility Functions
def setup_logging(name='commission_calculator'):
    """Configure logging with both file and console handlers."""
    logger.setLevel(logging.DEBUG)
    
    # Clear any existing handlers
    logger.handlers = []
    
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = f'{name}_{current_time}.log'
    
    # File handler - Debug level with detailed formatting
    fh = logging.FileHandler(log_filename)
    fh.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    fh.setFormatter(file_formatter)
    
    # Console handler - Info level with simpler formatting
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(message)s')
    ch.setFormatter(console_formatter)
    
    # Add console filter for commission calculator
    if name == 'commission_calculator':
        class ConsoleFilter(logging.Filter):
            def filter(self, record):
                return record.levelno == logging.INFO and not any(x in record.msg for x in [
                    'Department Summary',
                    'Completed:',
                    'Sales:',
                    'Total:',
                    'Spiffs:',
                    'Commissionable Revenue:',
                    'Commission:',
                    'Processing technician'
                ])
        ch.addFilter(ConsoleFilter())
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

def create_output_directory(base_path: str, start_of_week: datetime, end_of_week: datetime, logger: logging.Logger) -> str:
    """Create a directory for the output files with a name based on the week range."""
    # Format week range for folder name
    folder_name = f"Commission Output {start_of_week.strftime('%m_%d_%y')}-{end_of_week.strftime('%m_%d_%y')}"
    output_dir = os.path.join(base_path, folder_name)
    
    try:
        # Create the directory
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Created output directory: {output_dir}")
        return output_dir
    except Exception as e:
        logger.error(f"Error creating output directory: {str(e)}")
        raise


def autofit_columns(worksheet):
    """Autofit column widths in Excel worksheet."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def format_currency(amount):
    """Format number as currency string."""
    try:
        if isinstance(amount, str):
            amount = float(amount.replace('$', '').replace(',', ''))
        return f"${amount:,.2f}"
    except (ValueError, TypeError):
        return "$0.00"


def extract_subdepartment_code(business_unit):
    """Extract specific two-digit subdepartment code from business unit."""
    try:
        unit_str = str(business_unit).split('-')[0].strip()
        subdept = ''.join(filter(str.isdigit, unit_str))
        return subdept[:2] if len(subdept) >= 2 else '00'
    except:
        return '00'

def extract_department_number(business_unit):
    """Extract department number from business unit."""
    try:
        unit_str = str(business_unit).split('-')[0].strip()
        return int(''.join(filter(str.isdigit, unit_str)))
    except:
        return 0

def get_department_with_code(dept_num):
    """Get department name with code range."""
    if 20 <= dept_num <= 29:
        return 'HVAC (20-29)'
    elif 30 <= dept_num <= 39:
        return 'Plumbing (30-39)'
    elif 40 <= dept_num <= 49:
        return 'Electric (40-49)'
    return 'Unknown (0)'

def get_department_from_number(dept_num):
    """Get department name from number."""
    if 20 <= dept_num <= 29:
        return 'HVAC'
    elif 30 <= dept_num <= 39:
        return 'Plumbing'
    elif 40 <= dept_num <= 49:
        return 'Electric'
    return 'Unknown'

def extract_dept_code(memo: str) -> Optional[str]:
    """Extract department code from memo string."""
    if not memo:
        return None
    memo = str(memo).strip()
    match = re.match(r'^\s*(\d{2})(?:\s*-|\s+|$)', memo)
    return match.group(1) if match else None

def format_threshold_scale(thresholds):
    """Format threshold scale for display."""
    return (f"2%: ${thresholds[0]:,.0f} | "
            f"3%: ${thresholds[1]:,.0f} | "
            f"4%: ${thresholds[2]:,.0f} | "
            f"5%: ${thresholds[3]:,.0f}")

def extract_department_range(business_unit, logger):
    """Extract department range from business unit."""
    try:
        unit_number = int(''.join(filter(str.isdigit, business_unit)))
        base = (unit_number // 10) * 10
        if 20 <= base <= 20:
            return range(20, 30)
        elif 30 <= base <= 30:
            return range(30, 40)
        elif 40 <= base <= 40:
            return range(40, 50)
        return range(0, 0)
    except (ValueError, TypeError) as e:
        logger.warning(f"Could not extract department range from business unit: {business_unit}")
        return range(0, 0)


def is_same_department(source_unit, target_unit, logger=None):
    """Check if two business units are in the same department."""
    source_range = extract_department_range(source_unit, logger)
    target_range = extract_department_range(target_unit, logger)
    
    return (len(source_range) > 0 and 
            len(target_range) > 0 and 
            source_range.start == target_range.start)


def get_valid_tgls(file_path: str, tech_name: str) -> List[dict]:
    """Get valid TGLs for a technician."""
    try:
        tgl_df = pd.read_excel(file_path, sheet_name='Sheet1_TGL')
        logger.debug(f"Processing TGLs for {tech_name} from Sheet1_TGL")
        logger.debug(f"Available columns: {tgl_df.columns.tolist()}")
        
        valid_tgls = []
        tech_tgls = tgl_df[
            (tgl_df['Lead Generated By'] == tech_name) & 
            (tgl_df['Status'] == 'Completed')
        ]
        
        for _, tgl in tech_tgls.iterrows():
            source_unit = str(tgl['Business Unit'])
            target_unit = str(tgl['Lead Generated from Business Unit'])
            
            if is_same_department(source_unit, target_unit, logger):
                valid_tgls.append({
                    'job_number': tgl.get('Job #', 'N/A'),
                    'status': tgl['Status'],
                    'business_unit': source_unit,
                    'target_unit': target_unit,
                    'created_date': tgl['Created Date']
                })
                logger.debug(f"Valid TGL found - Job #: {tgl.get('Job #', 'N/A')}, "
                           f"From: {source_unit} To: {target_unit}")
        
        return valid_tgls
        
    except Exception as e:
        logger.error(f"Error processing TGL data for {tech_name}: {str(e)}")
        return []

def get_subdepartment_spiffs(file_path: str, tech_name: str) -> dict[str, float]:
    """
    Get spiffs broken down by subdepartment for display purposes only.
    """
    try:
        spiffs_df = pd.read_excel(file_path, sheet_name='Direct Payroll Adjustments')
        tech_spiffs = spiffs_df[spiffs_df['Technician'] == tech_name]
        
        # Initialize subdepartment totals
        subdepartment_spiffs = {
            code: 0 for code in ['20', '21', '22', '24', '25', '27', 
                               '30', '31', '33', '34', 
                               '40', '41', '42']
        }
        
        for _, spiff in tech_spiffs.iterrows():
            try:
                if pd.isna(spiff['Amount']) or pd.isna(spiff['Memo']):
                    continue
                    
                amount = float(str(spiff['Amount']).replace('$', '').replace(',', '').strip())
                if amount <= 0:  # Only filter out negative amounts
                    continue
                    
                memo = str(spiff['Memo']).strip()
                
                # Extract subdepartment code
                if not memo[:2].isdigit():
                    continue
                    
                subdept = memo[:2]
                if subdept in subdepartment_spiffs:
                    subdepartment_spiffs[subdept] += amount
                    
            except ValueError:
                continue
        
        return subdepartment_spiffs
        
    except Exception as e:
        logger.error(f"Error getting subdepartment spiffs for {tech_name}: {str(e)}")
        return {code: 0 for code in ['20', '21', '22', '24', '25', '27', 
                                   '30', '31', '33', '34', 
                                   '40', '41', '42']}

def get_spiffs_total(file_path: str, tech_name: str) -> tuple[float, dict[str, float]]:
    try:
        spiffs_df = pd.read_excel(file_path, sheet_name='Direct Payroll Adjustments')
        tech_spiffs = spiffs_df[spiffs_df['Technician'] == tech_name]
        
        department_spiffs = {
            'HVAC': 0,
            'Plumbing': 0,
            'Electric': 0
        }
        
        for idx, spiff in tech_spiffs.iterrows():
            try:
                # Skip if amount is missing
                if pd.isna(spiff['Amount']):
                    continue
                    
                # Convert amount to float
                amount = float(str(spiff['Amount']).replace('$', '').replace(',', '').strip())
                
                # Skip negative amounts and zero
                if amount <= 0:
                    continue
                    
                # Skip if memo is missing
                if pd.isna(spiff['Memo']):
                    continue
                    
                memo = str(spiff['Memo']).strip()
                
                # Skip if memo doesn't start with department number
                if not memo[:2].isdigit():
                    logger.warning(f"Invalid memo format - must start with department number. Row {idx + 2}: {memo}")
                    continue
                
                # Get department number and validate
                dept_num = int(memo[:2])
                if not (20 <= dept_num <= 29 or 30 <= dept_num <= 39 or 40 <= dept_num <= 49):
                    logger.warning(f"Invalid department number in memo (must be 20-29, 30-39, or 40-49). Row {idx + 2}: {memo}")
                    continue
                
                # Add amount to appropriate department total
                if 20 <= dept_num <= 29:
                    department_spiffs['HVAC'] += amount
                elif 30 <= dept_num <= 39:
                    department_spiffs['Plumbing'] += amount
                elif 40 <= dept_num <= 49:
                    department_spiffs['Electric'] += amount
                    
                logger.debug(f"Added positive spiff: ${amount:,.2f} to dept {dept_num}")
                
            except ValueError as e:
                logger.warning(f"Error processing spiff entry in row {idx + 2}: {str(e)}")
                continue
        
        spiffs_total = sum(department_spiffs.values())
        return spiffs_total, department_spiffs
        
    except Exception as e:
        logger.error(f"Error processing spiffs data for {tech_name}: {str(e)}")
        raise

def get_excused_hours(file_path: str, base_date: datetime, sheet_name: str = '2024') -> Dict[str, int]:
    """Get excused hours from time off sheet."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        start_of_week = base_date - timedelta(days=base_date.weekday())
        end_of_week = start_of_week + timedelta(days=4)
        
        def format_date(date):
            day = date.day
            if 4 <= day <= 20 or 24 <= day <= 30:
                suffix = "th"
            else:
                suffix = ["st", "nd", "rd"][day % 10 - 1]
            return f"{date.strftime('%B')} {day}{suffix}"
        
        target_week = f"{format_date(start_of_week)} - {format_date(end_of_week)}"
        logger.debug(f"Looking for week range: {target_week}")
        
        start_col = None
        for row_idx in [0, 1]:
            for col_idx, header in enumerate(df.iloc[row_idx]):
                if isinstance(header, str) and target_week.strip() == header.strip():
                    start_col = col_idx
                    logger.debug(f"Found target week in row {row_idx}, column {col_idx}")
                    break
            if start_col is not None:
                break
        
        if start_col is None:
            logger.warning(f"Week range '{target_week}' not found in time off sheet")
            return {}
        
        week_columns = list(range(start_col, start_col + 5))
        hours_summary = {}
        
        for index in range(2, len(df)):
            row = df.iloc[index]
            technician_name = row[0]
            
            if pd.isna(technician_name) or str(technician_name).strip() == '':
                continue
            
            technician_name = str(technician_name).strip()
            total_hours = 0
            
            for col in week_columns:
                try:
                    cell_value = str(row[col]).strip().lower()
                    if cell_value in ['x', 'r', 'v']:
                        total_hours += 8
                        logger.debug(f"Found time off marker '{cell_value}' for {technician_name} in column {col}")
                except Exception as e:
                    logger.warning(f"Error processing cell for {technician_name} in column {col}: {str(e)}")
                    continue
            
            if total_hours > 0:
                hours_summary[technician_name] = total_hours
                logger.debug(f"Found {total_hours} excused hours for {technician_name}")
        
        return hours_summary
        
    except Exception as e:
        logger.error(f"Error analyzing time off data: {str(e)}")
        return {}

def calculate_box_metrics(data: pd.DataFrame, tech_name: str, base_date: datetime) -> Tuple[float, float, float, Dict[str, Dict[str, float]]]:
    """Calculate Box A (CJR), Box B (TSIS), and Box C (Total) metrics with subdepartment breakdowns."""
    subdept_breakdown = {
        'completed': {code: 0.0 for code in SUBDEPARTMENT_MAP.keys()},
        'sales': {code: 0.0 for code in SUBDEPARTMENT_MAP.keys()},
        'total': {code: 0.0 for code in SUBDEPARTMENT_MAP.keys()}
    }

    # Calculate week range
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    logger.debug(f"\nCalculating metrics for {tech_name} for week {start_of_week.strftime('%Y-%m-%d')} to {end_of_week.strftime('%Y-%m-%d')}")

    # Ensure datetime format
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])

    # Get all jobs related to the technician in any capacity (primary or sold by)
    relevant_jobs = data[(data['Primary Technician'] == tech_name) | (data['Sold By'] == tech_name)]
    total_relevant_jobs = len(relevant_jobs)
    
    logger.debug(f"\nFound {total_relevant_jobs} total jobs related to {tech_name}")
    logger.debug("="*80)
    
    # Filter and log primary jobs within date range
    primary_jobs = relevant_jobs[
        (relevant_jobs['Primary Technician'] == tech_name) &
        (relevant_jobs['Invoice Date'].dt.date >= start_of_week.date()) &
        (relevant_jobs['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    logger.debug("\nCOMPLETED JOBS (Box A - CJR):")
    logger.debug("-" * 50)
    box_a = 0.0
    for _, job in primary_jobs.iterrows():
        if job.get('Opportunity', False):  # Only include opportunity jobs
            revenue = job['Jobs Total Revenue'] or 0
            box_a += revenue
            
            # Calculate subdepartment breakdowns for completed jobs
            subdept = extract_subdepartment_code(job.get('Business Unit', ''))
            subdept_breakdown['completed'][subdept] += revenue
            subdept_breakdown['total'][subdept] += revenue
            
            # Log each completed job's details
            logger.debug(f"Invoice #{job.get('Invoice #', 'N/A')} - {job['Invoice Date'].strftime('%m/%d/%y')}")
            logger.debug(f"Customer: {job.get('Customer Name', 'Unknown')}")
            logger.debug(f"Business Unit: {job.get('Business Unit', 'Unknown')}")
            logger.debug(f"Revenue: ${revenue:,.2f}")
            if pd.notna(job.get('GP')):
                logger.debug(f"GP: ${job.get('GP', 0):,.2f}")
            logger.debug(f"Opportunity: {'Yes' if job.get('Opportunity', False) else 'No'}")
            logger.debug("-" * 30)
    
    logger.debug(f"\nTotal Box A (CJR): ${box_a:,.2f}")
    logger.debug("="*80)
    
    # Filter and log sold jobs within date range
    sold_jobs = relevant_jobs[
        (relevant_jobs['Sold By'] == tech_name) &
        (relevant_jobs['Primary Technician'] != tech_name) &
        (relevant_jobs['Invoice Date'].dt.date >= start_of_week.date()) &
        (relevant_jobs['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    logger.debug("\nSOLD JOBS (Box B - TSIS):")
    logger.debug("-" * 50)
    box_b = 0.0
    for _, job in sold_jobs.iterrows():
        revenue = job['Jobs Total Revenue'] or 0
        box_b += revenue
        
        # Calculate subdepartment breakdowns for sold jobs
        subdept = extract_subdepartment_code(job.get('Business Unit', ''))
        subdept_breakdown['sales'][subdept] += revenue
        subdept_breakdown['total'][subdept] += revenue
        
        # Log each sold job's details
        logger.debug(f"Invoice #{job.get('Invoice #', 'N/A')} - {job['Invoice Date'].strftime('%m/%d/%y')}")
        logger.debug(f"Customer: {job.get('Customer Name', 'Unknown')}")
        logger.debug(f"Business Unit: {job.get('Business Unit', 'Unknown')}")
        logger.debug(f"Revenue: ${revenue:,.2f}")
        logger.debug(f"Primary Tech: {job.get('Primary Technician', 'Unknown')}")
        if pd.notna(job.get('GP')):
            logger.debug(f"GP: ${job.get('GP', 0):,.2f}")
        logger.debug(f"Opportunity: {'Yes' if job.get('Opportunity', False) else 'No'}")
        logger.debug("-" * 30)
    
    logger.debug(f"\nTotal Box B (TSIS): ${box_b:,.2f}")
    logger.debug("="*80)

    included_count = len(primary_jobs) + len(sold_jobs)
    skipped_count = total_relevant_jobs - included_count
    box_c = box_a + box_b
    
    logger.debug("\nSUMMARY:")
    logger.debug(f"Total Jobs Found: {total_relevant_jobs}")
    logger.debug(f"Jobs Included: {included_count}")
    logger.debug(f"Jobs Skipped: {skipped_count}")
    logger.debug(f"Box A (CJR): ${box_a:,.2f}")
    logger.debug(f"Box B (TSIS): ${box_b:,.2f}")
    logger.debug(f"Box C (Total): ${box_c:,.2f}")
    
    # Log summary of skipped jobs if any
    if skipped_count > 0:
        logger.info(f"Note: {skipped_count} jobs were skipped because they did not fall within the selected week")
        skipped_jobs = relevant_jobs[
            ~(
                ((relevant_jobs['Primary Technician'] == tech_name) |
                (relevant_jobs['Sold By'] == tech_name)) &
                (relevant_jobs['Invoice Date'].dt.date >= start_of_week.date()) &
                (relevant_jobs['Invoice Date'].dt.date <= end_of_week.date())
            )
        ]
        for _, job in skipped_jobs.iterrows():
            logger.debug(f"\nSkipped Job Details:")
            logger.debug(f"Invoice #{job.get('Invoice #', 'N/A')} - {job['Invoice Date'].strftime('%m/%d/%y')}")
            logger.debug(f"Customer: {job.get('Customer Name', 'Unknown')}")
            logger.debug(f"Business Unit: {job.get('Business Unit', 'Unknown')}")
            logger.debug(f"Revenue: ${job.get('Jobs Total Revenue', 0):,.2f}")
            logger.debug(f"Primary Tech: {job.get('Primary Technician', 'Unknown')}")
            logger.debug(f"Sold By: {job.get('Sold By', 'Unknown')}")
        
    return box_a, box_b, box_c, subdept_breakdown

def calculate_percentages(box_a: float, box_c: float) -> Tuple[int, int]:
    """Calculate Service Completion and Install Contribution percentages."""
    if box_c == 0:
        return 0, 0
        
    raw_scp = (box_a / box_c * 100)
    raw_icp = 100 - raw_scp
    
    raw_scp = max(0, raw_scp)
    raw_icp = max(0, raw_icp)
    
    total = raw_scp + raw_icp
    if total != 0:
        raw_scp = (raw_scp / total) * 100
        raw_icp = (raw_icp / total) * 100
        
    scp = round(raw_scp / 10) * 10
    icp = round(raw_icp / 10) * 10
    
    if scp + icp != 100:
        if scp > icp:
            scp = 100
            icp = 0
        else:
            scp = 0
            icp = 100
    
    return int(scp), int(icp)

def calculate_average_ticket_value(data: pd.DataFrame, tech_name: str, box_a: float, box_b: float, base_date: datetime, logger: logging.Logger) -> Dict[str, float]:
    """Calculate average ticket value using total revenue divided by opportunity count."""
    avg_tickets = {'overall': 0.0}
    
    # Calculate week range
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    logger.debug(f"\nCALCULATING AVERAGE TICKET VALUE FOR {tech_name.upper()}")
    logger.debug("=" * 80)
    logger.debug(f"Week Range: {start_of_week.strftime('%m/%d/%y')} to {end_of_week.strftime('%m/%d/%y')}")
    
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])

    # Get completed jobs within date range
    completed_jobs = data[
        (data['Primary Technician'] == tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    # Get count of opportunity jobs and log details
    opportunity_jobs = completed_jobs[completed_jobs['Opportunity'] == True]
    opportunity_count = len(opportunity_jobs)
    
    logger.debug("\nOPPORTUNITY JOBS BREAKDOWN:")
    logger.debug("-" * 50)
    
    total_revenue = box_a + box_b
    
    # Track department totals
    dept_totals = {
        'HVAC': {'count': 0, 'revenue': 0.0},
        'Plumbing': {'count': 0, 'revenue': 0.0},
        'Electric': {'count': 0, 'revenue': 0.0}
    }
    
    for _, job in opportunity_jobs.iterrows():
        revenue = job.get('Jobs Total Revenue', 0) or 0
        dept_num = extract_department_number(str(job.get('Business Unit', '')))
        dept = get_department_from_number(dept_num)
        
        # Update department totals
        if dept in dept_totals:
            dept_totals[dept]['count'] += 1
            dept_totals[dept]['revenue'] += revenue
        
        logger.debug(f"\nInvoice #{job.get('Invoice #', 'N/A')} - {job['Invoice Date'].strftime('%m/%d/%y')}")
        logger.debug(f"Customer: {job.get('Customer Name', 'Unknown')}")
        logger.debug(f"Department: {dept}")
        logger.debug(f"Business Unit: {job.get('Business Unit', 'Unknown')}")
        logger.debug(f"Revenue: ${revenue:,.2f}")
        if pd.notna(job.get('GP')):
            logger.debug(f"GP: ${job.get('GP', 0):,.2f}")
    
    avg_ticket = round(total_revenue / opportunity_count, 2) if opportunity_count > 0 else 0
    avg_tickets['overall'] = avg_ticket
    
    logger.debug("\nDEPARTMENT SUMMARY:")
    logger.debug("-" * 50)
    for dept, totals in dept_totals.items():
        if totals['count'] > 0:
            dept_avg = totals['revenue'] / totals['count']
            logger.debug(f"\n{dept}:")
            logger.debug(f"Number of Opportunities: {totals['count']}")
            logger.debug(f"Total Revenue: ${totals['revenue']:,.2f}")
            logger.debug(f"Average Ticket: ${dept_avg:,.2f}")
    
    logger.debug("\nOVERALL SUMMARY:")
    logger.debug("-" * 50)
    logger.debug(f"Box A (CJR): ${box_a:,.2f}") 
    logger.debug(f"Box B (TSIS): ${box_b:,.2f}")
    logger.debug(f"Total Revenue (CJR + TSIS): ${total_revenue:,.2f}")
    logger.debug(f"Total Opportunity Count: {opportunity_count}")
    logger.debug(f"Overall Average Ticket: ${avg_ticket:,.2f}")
    
    # Log non-opportunity jobs if any exist
    non_opp_jobs = completed_jobs[completed_jobs['Opportunity'] == False]
    if not non_opp_jobs.empty:
        logger.debug("\nNON-OPPORTUNITY JOBS (Not Included in Average):")
        logger.debug("-" * 50)
        for _, job in non_opp_jobs.iterrows():
            logger.debug(f"\nInvoice #{job.get('Invoice #', 'N/A')} - {job['Invoice Date'].strftime('%m/%d/%y')}")
            logger.debug(f"Customer: {job.get('Customer Name', 'Unknown')}")
            logger.debug(f"Business Unit: {job.get('Business Unit', 'Unknown')}")
            logger.debug(f"Revenue: ${job.get('Jobs Total Revenue', 0):,.2f}")
    
    return avg_tickets

def format_department_revenue(revenue_data: Dict[str, Dict[str, float]], 
                            commission_rate: float,
                            department_spiffs: Dict[str, float],
                            subdept_breakdown: Dict[str, Dict[str, float]],
                            subdepartment_spiffs: Dict[str, float]) -> Dict[str, str]:
    """Format department and subdepartment revenue data into strings for Excel output."""
    formatted = {}
    
    # Helper function to convert string amounts to float
    def parse_amount(amount_str: str) -> float:
        try:
            return float(amount_str.replace('$', '').replace(',', ''))
        except (ValueError, AttributeError):
            return 0.0
    
    # Initialize department commission totals
    dept_commission_totals = {
        'HVAC': 0.0,
        'Plumbing': 0.0,
        'Electric': 0.0
    }
    
    # Format subdepartment totals and accumulate department commissions
    for subdept_code in ['20', '21', '22', '24', '25', '27', '30', '31', '33', '34', '40', '41', '42']:
        completed = subdept_breakdown['completed'].get(subdept_code, 0)
        sales = subdept_breakdown['sales'].get(subdept_code, 0)
        spiffs = subdepartment_spiffs.get(subdept_code, 0)
        total = subdept_breakdown['total'].get(subdept_code, 0)
        
        formatted[f"{subdept_code} Revenue"] = f"${completed:,.2f}"
        formatted[f"{subdept_code} Sales"] = f"${sales:,.2f}"
        formatted[f"{subdept_code} Spiffs"] = f"${spiffs:,.2f}"
        formatted[f"{subdept_code} Total"] = f"${total:,.2f}"
        
        # Calculate commission for this subdepartment, ensuring it's never negative
        calc_total = max(0, (completed + sales - spiffs) * commission_rate)
        formatted[f"{subdept_code} Commission"] = f"${calc_total:,.2f}"
        
        # Add to department totals based on subdepartment code
        if subdept_code.startswith('2'):
            dept_commission_totals['HVAC'] += calc_total
        elif subdept_code.startswith('3'):
            dept_commission_totals['Plumbing'] += calc_total
        elif subdept_code.startswith('4'):
            dept_commission_totals['Electric'] += calc_total
    
    # Format main department totals
    for dept in ['HVAC', 'Plumbing', 'Electric']:
        completed = revenue_data['completed'][dept]
        sales = revenue_data['sales'][dept]
        combined = revenue_data['combined'][dept]
        dept_spiffs = department_spiffs[dept]
        adjusted_combined = max(0, combined - dept_spiffs)
        
        formatted[f"{dept} Revenue"] = f"${completed:,.2f}"
        formatted[f"{dept} Sales"] = f"${sales:,.2f}"
        formatted[f"{dept} Spiffs"] = f"${dept_spiffs:,.2f}"
        formatted[f"{dept} Total"] = f"${adjusted_combined:,.2f}"
        # Use the summed commission from subdepartments
        formatted[f"{dept} Commission"] = f"${dept_commission_totals[dept]:,.2f}"
    
    return formatted

def calculate_department_revenue(data: pd.DataFrame, tech_name: str, base_date: datetime) -> Dict[str, Dict[str, float]]:
    """Calculate department revenue using the same date filtering."""
    # Calculate week range
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    revenue_by_dept = {
        'completed': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0},
        'sales': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0},
        'combined': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0}
    }
    
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])

    # Get completed jobs within date range
    completed_jobs = data[
        (data['Primary Technician'] == tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    # Process completed jobs by department
    for dept in ['HVAC', 'Plumbing', 'Electric']:
        dept_completed_jobs = completed_jobs[
            completed_jobs['Business Unit'].apply(
                lambda x: get_department_from_number(extract_department_number(str(x))) == dept
            )
        ]
        
        dept_total = 0.0
        for _, job in dept_completed_jobs.iterrows():
            if job.get('Opportunity', False):
                revenue = job.get('Jobs Total Revenue', 0) or 0
                dept_total += revenue
                
        revenue_by_dept['completed'][dept] = dept_total
        revenue_by_dept['combined'][dept] = dept_total
    
    # Get sales within date range
    sold_jobs = data[
        (data['Sold By'] == tech_name) & 
        (data['Primary Technician'] != tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    # Process sold jobs by department
    for dept in ['HVAC', 'Plumbing', 'Electric']:
        dept_sold_jobs = sold_jobs[
            sold_jobs['Business Unit'].apply(
                lambda x: get_department_from_number(extract_department_number(str(x))) == dept
            )
        ]
        
        dept_total = 0.0
        for _, job in dept_sold_jobs.iterrows():
            revenue = job.get('Jobs Total Revenue', 0) or 0
            dept_total += revenue
            
        revenue_by_dept['sales'][dept] = dept_total
        revenue_by_dept['combined'][dept] += dept_total  # Add to existing completed revenue
    
    return revenue_by_dept

def get_commission_rate(total_revenue: float, flipped_percent: float, department: str, 
                       excused_hours: int, tgl_reduction: float, avg_ticket_value: float) -> Tuple[float, list, list]:
    """Calculate commission rate and thresholds based on revenue and department."""
    logger.debug("\nDETAILED THRESHOLD CALCULATION")
    logger.debug("=" * 80)
    
    # Round flipped percent to nearest 10
    flipped_percent = min(100, max(0, int(round(flipped_percent / 10) * 10)))
    logger.debug(f"Install Contribution Percentage (ICP): {flipped_percent}%")
    
    # Get base thresholds
    if department in ['Electric', 'Plumbing']:
        thresholds = PLUMBING_ELECTRICAL_THRESHOLDS
        logger.debug(f"Using {department} threshold table")
    else:
        thresholds = HVAC_THRESHOLDS
        logger.debug("Using HVAC threshold table")
    
    tier_thresholds = thresholds[flipped_percent].copy()
    logger.debug(f"\nBase thresholds for {department} at {flipped_percent}% ICP:")
    logger.debug(f"2% Tier: ${tier_thresholds[0]:,.2f}")
    logger.debug(f"3% Tier: ${tier_thresholds[1]:,.2f}")
    logger.debug(f"4% Tier: ${tier_thresholds[2]:,.2f}")
    logger.debug(f"5% Tier: ${tier_thresholds[3]:,.2f}")
    
    # Calculate time off reduction
    days_off = min(5, excused_hours / 8)
    reduction_factor = max(0, 1 - (0.20 * days_off))
    logger.debug(f"\nTime Off Adjustment:")
    logger.debug(f"Excused Hours: {excused_hours}")
    logger.debug(f"Days Off: {days_off}")
    logger.debug(f"Reduction Factor: {reduction_factor:.2f} (Reduces thresholds by {(1-reduction_factor)*100:.1f}%)")
    
    # Apply time off reduction
    time_off_adjusted = [threshold * reduction_factor for threshold in tier_thresholds]
    logger.debug("\nThresholds after time off adjustment:")
    logger.debug(f"2% Tier: ${time_off_adjusted[0]:,.2f}")
    logger.debug(f"3% Tier: ${time_off_adjusted[1]:,.2f}")
    logger.debug(f"4% Tier: ${time_off_adjusted[2]:,.2f}")
    logger.debug(f"5% Tier: ${time_off_adjusted[3]:,.2f}")
    
   # Calculate and apply TGL reduction
    tgl_count = int(tgl_reduction / avg_ticket_value) if avg_ticket_value > 0 else 0
    logger.debug("\nTGL Reduction Calculation:")
    logger.debug(f"Number of Valid TGLs: {tgl_count}")
    logger.debug(f"Average Ticket Value: ${avg_ticket_value:,.2f}")
    logger.debug(f"TGL Credit = {tgl_count} TGLs × ${avg_ticket_value:,.2f} = ${tgl_reduction:,.2f}")
    
    # Show threshold reduction details for each tier
    logger.debug("\nApplying TGL reduction to each threshold tier:")
    for i, threshold in enumerate(time_off_adjusted):
        tier_percent = (i + 2)  # 2%, 3%, 4%, 5%
        adjusted_value = max(0, threshold - tgl_reduction)
        reduction = threshold - adjusted_value
        logger.debug(f"{tier_percent}% Tier: ${threshold:,.2f} - ${tgl_reduction:,.2f} = ${adjusted_value:,.2f}")
        logger.debug(f"  • Original threshold: ${threshold:,.2f}")
        logger.debug(f"  • TGL reduction: ${tgl_reduction:,.2f}")
        logger.debug(f"  • Final threshold: ${adjusted_value:,.2f}")
    
    # Apply TGL reduction
    adjusted_thresholds = [max(0, threshold - tgl_reduction) for threshold in time_off_adjusted]
    logger.debug("\nFinal thresholds after TGL reduction:")
    logger.debug(f"2% Tier: ${adjusted_thresholds[0]:,.2f}")
    logger.debug(f"3% Tier: ${adjusted_thresholds[1]:,.2f}")
    logger.debug(f"4% Tier: ${adjusted_thresholds[2]:,.2f}")
    logger.debug(f"5% Tier: ${adjusted_thresholds[3]:,.2f}")

    # Determine commission rate based on highest threshold met
    logger.debug("\nRevenue vs Threshold Comparison:")
    logger.debug(f"Total Revenue: ${total_revenue:,.2f}")
    
    if total_revenue >= adjusted_thresholds[3]:
        rate = 0.05
        logger.debug(f"Revenue exceeds 5% tier (${adjusted_thresholds[3]:,.2f})")
    elif total_revenue >= adjusted_thresholds[2]:
        rate = 0.04
        logger.debug(f"Revenue exceeds 4% tier (${adjusted_thresholds[2]:,.2f})")
    elif total_revenue >= adjusted_thresholds[1]:
        rate = 0.03
        logger.debug(f"Revenue exceeds 3% tier (${adjusted_thresholds[1]:,.2f})")
    elif total_revenue >= adjusted_thresholds[0]:
        rate = 0.02
        logger.debug(f"Revenue exceeds 2% tier (${adjusted_thresholds[0]:,.2f})")
    else:
        rate = 0
        logger.debug("Revenue did not meet minimum threshold")
        logger.debug(f"Needed ${adjusted_thresholds[0]:,.2f} for 2% tier, short by ${adjusted_thresholds[0] - total_revenue:,.2f}")
    
    logger.debug(f"\nFinal Commission Rate: {rate*100}%")
    return rate, adjusted_thresholds, tier_thresholds

def process_commission_calculations(data: pd.DataFrame, tech_data: pd.DataFrame, 
                                 file_path: str, base_date: datetime,
                                 excused_hours_dict: Dict[str, int]) -> pd.DataFrame:
    results = []
    
    # Filter tech_data to only include service technicians
    service_techs = tech_data[
        (~tech_data['Name'].isin(EXCLUDED_TECHS)) &
        (tech_data['Technician Business Unit'].apply(
            lambda x: determine_tech_type(x) == 'SERVICE'
        ))
    ]['Name'].tolist()
    
    # Create a mapping of technician names to their exact business units
    tech_dept_map = dict(zip(tech_data['Name'], tech_data['Technician Business Unit']))

    for tech_name in service_techs:
        logger.info(f"\nProcessing technician: {tech_name}")
        # Get badge ID from tech_data
        badge_id = tech_data.loc[tech_data['Name'] == tech_name, 'Badge ID'].iloc[0]
        
        # Calculate metrics
        box_a, box_b, box_c, subdept_breakdown = calculate_box_metrics(data, tech_name, base_date)
        scp, icp = calculate_percentages(box_a, box_c)
        
        dept_revenue = calculate_department_revenue(data, tech_name, base_date)
        
        # Get department spiffs (used for actual calculations)
        spiffs_total, department_spiffs = get_spiffs_total(file_path, tech_name)
        
        # Get subdepartment spiffs (for display only)
        subdepartment_spiffs = get_subdepartment_spiffs(file_path, tech_name)
        
        valid_tgls = get_valid_tgls(file_path, tech_name)
        
        avg_tickets = calculate_average_ticket_value(data, tech_name, box_a, box_b, base_date, logger)
        default_ticket = 0.0
        avg_ticket_value = avg_tickets.get('overall', default_ticket) if avg_tickets else default_ticket
        
        # Get exact business unit from mapping
        business_unit = tech_dept_map.get(tech_name, 'Unknown')
        
        # Extract department for commission calculations only
        dept_num = extract_department_number(business_unit)
        department = get_department_from_number(dept_num)

        tgl_reduction = avg_ticket_value * len(valid_tgls) if avg_ticket_value > 0 else 0
        
        excused_hours = excused_hours_dict.get(tech_name, 0)
        
        commission_rate, adjusted_thresholds, base_thresholds = get_commission_rate(
            box_c, icp, department, excused_hours, tgl_reduction, avg_ticket_value
        )
        
        base_threshold_scale = format_threshold_scale(base_thresholds)
        adjusted_threshold_scale = format_threshold_scale(adjusted_thresholds)
        
        formatted_dept_data = format_department_revenue(
            dept_revenue,
            commission_rate,
            department_spiffs,
            subdept_breakdown,
            subdepartment_spiffs
        )
        
        # Calculate final commission using original logic (department level spiffs)
        commissionable_revenue = box_c - spiffs_total
        final_commission = commissionable_revenue * commission_rate
        
        result = {
            'Badge ID': badge_id,
            'Technician': tech_name,
            'Main Dept': business_unit,  # Use exact business unit from Sheet1_Tech
            'Total Revenue': box_c,
            'Completed Job Revenue': box_a,
            'Tech-Sourced Install Sales': box_b,
            'Service Completion %': scp,
            'Install Contribution %': icp,
            'Excused Hours': excused_hours,
            'Spiffs': spiffs_total,
            'Valid TGLs': len(valid_tgls),
            'Avg Ticket $': avg_ticket_value,
            'TGL Threshold Reduction': tgl_reduction,
            'Base Threshold Scale': base_threshold_scale,
            'Adjusted Threshold Scale': adjusted_threshold_scale,
            'Commissionable Revenue': commissionable_revenue,
            'Commission Rate %': commission_rate * 100,
            'Total Commission': round(final_commission, 2),
            'Status': f"Qualified for {commission_rate*100}% tier" if commission_rate > 0 else "Did not qualify"
        }
        
        # Add formatted department data including display-only subdepartment spiffs
        result.update(formatted_dept_data)
        results.append(result)

    results_df = pd.DataFrame(results)
    
    # Ensure all columns are present in the correct order
    for col in COLUMN_ORDER:
        if col not in results_df.columns:
            results_df[col] = ''
    
    return results_df[COLUMN_ORDER]

def read_tech_department_data(file_path: str, logger: logging.Logger) -> pd.DataFrame:
    try:
        logger.debug(f"Reading technician department data from {file_path}")
        # Read the Excel file
        tech_df = pd.read_excel(file_path, sheet_name='Sheet1_Tech', dtype={'Payroll ID': str})
        
        # Remove rows where Name is numeric
        tech_df = tech_df[~tech_df['Name'].astype(str).str.isnumeric()]
        
        # Filter out excluded techs
        tech_df = tech_df[~tech_df['Name'].isin(EXCLUDED_TECHS)]
        
        # Format badge IDs
        tech_df['Badge ID'] = tech_df['Payroll ID'].apply(format_badge_id)
        
        logger.debug(f"Successfully loaded {len(tech_df)} technician records")
        return tech_df
        
    except Exception as e:
        logger.error(f"Error reading technician department data: {str(e)}")
        raise

def determine_pay_code(business_unit: str) -> Optional[str]:
    """Determine pay code based on business unit description."""
    if pd.isna(business_unit):
        return None
    
    bu_upper = str(business_unit).upper()
    if 'ADMINISTRATIVE' in bu_upper:
        return None
    elif 'SERVICE' in bu_upper:
        return 'PCM'
    else:
        return 'ICM'

def sum_spiffs_for_dept(spiffs_df: pd.DataFrame, tech_name: str, dept_code: str) -> float:
    """Sum spiffs for a specific technician and department code."""
    dept_spiffs = spiffs_df[
        (spiffs_df['Technician'] == tech_name) & 
        (spiffs_df['Memo'].apply(lambda x: extract_dept_code(str(x)) == dept_code)) &
        (spiffs_df['Amount'].apply(
            lambda x: float(str(x).replace('$', '').replace(',', '')) > 0 
            if pd.notnull(x) else False
        ))
    ]
    
    if dept_spiffs.empty:
        return 0.0
        
    amounts = dept_spiffs['Amount'].apply(
        lambda x: float(str(x).replace('$', '').replace(',', '')) if pd.notnull(x) else 0.0
    )
    
    return amounts.sum()

def process_paystats(output_dir: str, paystats_file: str, tech_data: pd.DataFrame, 
                    base_date: datetime, logger: logging.Logger) -> List[PayrollEntry]:
    """
    Process payroll entries for service technicians from paystats file with updated department logic.
    Now uses service department codes for PCM entries regardless of technician's home department.
    
    Args:
        output_dir (str): Directory containing output files
        paystats_file (str): Path to the paystats Excel file
        tech_data (pd.DataFrame): DataFrame containing technician information
        base_date (datetime): Base date for processing
        logger (logging.Logger): Logger instance
        
    Returns:
        List[PayrollEntry]: List of processed payroll entries
    """
    logger.info("Processing payroll entries from paystats file...")
    payroll_entries = []

    try:
        # Calculate week range and end date
        start_of_week = base_date - timedelta(days=base_date.weekday())
        week_end_date = start_of_week + timedelta(days=6)
        target_date = week_end_date.strftime('%m/%d/%Y')
        
        stats_df = pd.read_excel(paystats_file)
        adj_df = pd.read_excel(os.path.join(output_dir, 'combined_data.xlsx'), 
                             sheet_name='Direct Payroll Adjustments')

        # Filter to include only service technicians
        stats_df = stats_df[
            (~stats_df['Technician'].isin(EXCLUDED_TECHS)) &
            (stats_df['Technician'].isin(
                tech_data[
                    tech_data['Technician Business Unit'].apply(
                        lambda x: determine_tech_type(x) == 'SERVICE'
                    )
                ]['Name']
            ))
        ]

        # Process each technician's entries
        for _, row in stats_df.iterrows():
            tech_name = row['Technician']
            commission_rate = row['Commission Rate %'] / 100

            # Skip if no commission rate
            if commission_rate == 0:
                continue

            # Get technician info
            tech_info = tech_data[tech_data['Name'] == tech_name]
            if tech_info.empty:
                continue

            # Get badge ID (try both Badge ID and Payroll ID fields)
            badge_id = None
            if 'Badge ID' in tech_info.columns:
                badge_id = tech_info.iloc[0]['Badge ID'] if pd.notna(tech_info.iloc[0]['Badge ID']) else None
            if badge_id is None and 'Payroll ID' in tech_info.columns:
                badge_id = tech_info.iloc[0]['Payroll ID'] if pd.notna(tech_info.iloc[0]['Payroll ID']) else None
            
            if badge_id is None:
                logger.warning(f"Skipping {tech_name} - No valid Badge ID or Payroll ID found")
                continue
            
            # Process each subdepartment
            for dept_code in DEPARTMENT_CODES.keys():
                revenue_col = f"{dept_code} Revenue"
                sales_col = f"{dept_code} Sales"
                total_col = f"{dept_code} Total"
                
                if not all(col in row.index for col in [revenue_col, sales_col, total_col]):
                    continue
                
                # Parse amounts
                def parse_amount(val):
                    if pd.isna(val):
                        return 0.0
                    if isinstance(val, str):
                        return float(val.replace('$', '').replace(',', '').strip() or 0)
                    return float(val or 0)

                revenue = parse_amount(row[revenue_col])
                sales = parse_amount(row[sales_col])
                total = parse_amount(row[total_col])

                if revenue == 0 and sales == 0 and total == 0:
                    continue

                try:
                    # Get spiffs for this department
                    dept_spiffs = adj_df[
                        (adj_df['Technician'] == tech_name) & 
                        (adj_df['Memo'].apply(lambda x: extract_dept_code(str(x)) == dept_code))
                    ]
                    
                    # Sum only positive spiffs (negatives are handled separately)
                    positive_spiffs = sum(
                        float(str(amount).replace('$', '').replace(',', ''))
                        for amount in dept_spiffs['Amount']
                        if pd.notna(amount) and 
                        float(str(amount).replace('$', '').replace(',', '')) > 0
                    )
                    
                    # Adjust total by positive spiffs only
                    adjusted_amount = round(total - positive_spiffs, 2)
                    if adjusted_amount <= 0:
                        continue

                    final_amount = round(adjusted_amount * commission_rate, 2)
                    if final_amount <= 0:
                        continue

                    # Create PCM entry using the service department's main code
                    main_dept_code = get_service_department_code(dept_code)
                    entry = PayrollEntry(
                        company_code=COMPANY_CODE,
                        badge_id=badge_id,
                        date=target_date,
                        amount=final_amount,
                        pay_code='PCM',
                        dept=main_dept_code,
                        location_id=LOCATION_ID
                    )
                    payroll_entries.append(entry)
                    
                    logger.debug(f"Created PCM entry for {tech_name} in dept {dept_code}:")
                    logger.debug(f"  Total: ${total:,.2f}")
                    logger.debug(f"  Positive Spiffs: ${positive_spiffs:,.2f}")
                    logger.debug(f"  Adjusted Amount: ${adjusted_amount:,.2f}")
                    logger.debug(f"  Commission Rate: {commission_rate*100}%")
                    logger.debug(f"  Final Amount: ${final_amount:,.2f}")
                    logger.debug(f"  Department Code: {main_dept_code}")

                except Exception as e:
                    logger.warning(f"Error processing department {dept_code} for {tech_name}: {str(e)}")
                    continue

        return payroll_entries

    except Exception as e:
        logger.error(f"Error processing paystats file: {str(e)}")
        raise

def process_gp_entries(output_dir: str, tech_data: pd.DataFrame, base_date: datetime, logger: logging.Logger) -> List[PayrollEntry]:
    """Process GP entries from Invoices sheet, specifically for installers."""
    logger.info("Processing GP entries for installers from Invoices sheet...")
    payroll_entries = []

    try:
        # Calculate week end date (Sunday)
        start_of_week = base_date - timedelta(days=base_date.weekday())
        week_end_date = start_of_week + timedelta(days=6)
        target_date = week_end_date.strftime('%m/%d/%Y')  # Use week end date

        # Read invoices data - specifically from the "Invoices" sheet
        invoices_df = pd.read_excel(os.path.join(output_dir, 'combined_data.xlsx'), sheet_name='Invoices')
        logger.debug(f"Loaded {len(invoices_df)} invoice records from Invoices sheet")

        # Convert GP column from currency string to float
        def parse_currency(value):
            if pd.isna(value):
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            try:
                cleaned = str(value).replace('$', '').replace(',', '').strip()
                return float(cleaned) if cleaned else 0.0
            except (ValueError, TypeError):
                return 0.0

        # Convert GP and other currency columns
        currency_columns = ['GP', 'Total', 'Cost', 'Subtotal']
        for col in currency_columns:
            if col in invoices_df.columns:
                invoices_df[col] = invoices_df[col].apply(parse_currency)
        
        # Filter tech_data to only include installers
        install_techs = tech_data[
            (~tech_data['Name'].isin(EXCLUDED_TECHS)) &
            (tech_data['Technician Business Unit'].apply(
                lambda x: determine_tech_type(x) == 'INSTALL'
            ))
        ]
        
        logger.info(f"Processing GP for {len(install_techs)} installers")

        # Prepare columns for merge
        merge_columns = ['Name']
        if 'Badge ID' in install_techs.columns:
            merge_columns.append('Badge ID')
        if 'Payroll ID' in install_techs.columns:
            merge_columns.append('Payroll ID')
        merge_columns.append('Technician Business Unit')

        # Merge installer data with invoices
        merged_df = invoices_df.merge(
            install_techs[merge_columns],
            left_on='Technician',
            right_on='Name',
            how='inner'
        )
        
        logger.debug(f"Found {len(merged_df)} invoice records for installers")

        # Group GP values and filter out zeros and nulls
        group_columns = ['Name', 'Business Unit']
        if 'Badge ID' in merged_df.columns:
            group_columns.append('Badge ID')
        if 'Payroll ID' in merged_df.columns:
            group_columns.append('Payroll ID')

        grouped = merged_df.groupby(group_columns)['GP'].sum().reset_index()
        grouped = grouped[
            (grouped['GP'] != 0) & 
            (grouped['GP'].notna()) & 
            (grouped['GP'] > 0)
        ]
        
        logger.debug(f"Found {len(grouped)} valid GP entries after grouping")

        for _, row in grouped.iterrows():
            # Get badge ID (try both Badge ID and Payroll ID fields)
            badge_id = None
            if 'Badge ID' in row.index:
                badge_id = row['Badge ID'] if pd.notna(row['Badge ID']) else None
            if badge_id is None and 'Payroll ID' in row.index:
                badge_id = row['Payroll ID'] if pd.notna(row['Payroll ID']) else None
            
            if badge_id is None:
                logger.warning(f"Skipping entry for {row['Name']} - No valid Badge ID or Payroll ID found")
                continue

            if pd.isna(row['Business Unit']):
                logger.warning(f"Skipping entry for {row['Name']} due to missing Business Unit")
                continue

            try:
                # Extract department code from business unit
                business_unit = str(row['Business Unit']).upper()
                dept_codes = re.findall(r'\d{2}', business_unit)
                if not dept_codes:
                    logger.warning(f"Could not find department code in business unit: {business_unit}")
                    continue
                
                subdepartment_code = dept_codes[0]
                dept_code = DEPARTMENT_CODES.get(subdepartment_code, {}).get('code')
                
                if not dept_code:
                    logger.warning(f"Invalid department code {subdepartment_code} found in: {business_unit}")
                    continue

                gp_value = float(row['GP'])
                if gp_value <= 0:
                    logger.debug(f"Skipping non-positive GP value for {row['Name']}: ${gp_value}")
                    continue

                # Create payroll entry with standardized ICM code for installers
                # and using the week end date
                entry = PayrollEntry(
                    company_code=COMPANY_CODE,
                    badge_id=badge_id,
                    date=target_date,  # Using week end date
                    amount=gp_value,
                    pay_code='ICM',  # Standardized to ICM for all installer entries
                    dept=dept_code,
                    location_id=LOCATION_ID
                )
                payroll_entries.append(entry)
                logger.info(f"Created GP entry for installer {row['Name']}: ${gp_value:,.2f} in dept {dept_code}")

            except Exception as e:
                logger.error(f"Error processing GP entry for {row['Name']}: {str(e)}")
                continue

        logger.info(f"Successfully processed {len(payroll_entries)} GP entries for installers")
        return payroll_entries

    except Exception as e:
        logger.error(f"Error processing installer GP entries: {str(e)}")
        raise

def match_department_spiffs(adj_df: pd.DataFrame, logger: logging.Logger) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Match positive and negative spiffs within departments and track calculations."""
    logger.info("Processing and matching department spiffs...")
    
    # Filter out excluded techs
    adj_df = adj_df[~adj_df['Technician'].isin(EXCLUDED_TECHS)]
    
    tgl_entries = []
    matched_entries = []
    unmatched_negatives = []
    
    # First process TGLs (this part stays the same)
    tgl_rows = adj_df[adj_df['Memo'].str.contains('tgl', case=False, na=False)]
    for _, row in tgl_rows.iterrows():
        amount = float(str(row['Amount']).replace('$', '').replace(',', ''))
        memo = str(row['Memo']).strip()
        dept_code = memo[:2]
        
        tgl_entries.append({
            'Technician': row['Technician'],
            'Department': dept_code,
            'Memo': memo,
            'Amount': amount,
            'Type': 'TGL'
        })
    
    # Process spiffs by technician and department
    for tech_name, tech_group in adj_df.groupby('Technician'):
        # Use new helper function to get accumulated totals
        dept_totals = process_department_entries(tech_group)
        
        # Create entries for each department
        for dept_code, totals in dept_totals.items():
            pos_total = totals['positives']
            neg_total = totals['negatives']  # Already negative
            net_amount = pos_total + neg_total  # This will properly subtract negatives
            
            logger.debug(f"\nProcessing {tech_name} - Dept {dept_code}:")
            logger.debug(f"  Total Positives: ${pos_total:,.2f}")
            logger.debug(f"  Total Negatives: ${neg_total:,.2f}")
            logger.debug(f"  Net Amount: ${net_amount:,.2f}")
            
            matched_entries.append({
                'Technician': tech_name,
                'Department': f"{dept_code} - {DEPARTMENT_CODES.get(dept_code, {'desc': 'Unknown'})['desc']}",
                'Original Positive': pos_total,
                'Negative Amount': neg_total,
                'Net Amount': net_amount,
                'Calculation': f"Positive(${pos_total:,.2f}) + Negative(${neg_total:,.2f}) = ${net_amount:,.2f}",
                'Status': 'Fully Processed'
            })
            
            # Track significant negative balances for reporting
            if net_amount < 0:
                unmatched_negatives.append({
                    'Technician': tech_name,
                    'Department': f"{dept_code} - {DEPARTMENT_CODES.get(dept_code, {'desc': 'Unknown'})['desc']}",
                    'Original Positive': pos_total,
                    'Negative Amount': neg_total,
                    'Remaining Amount': net_amount,
                    'Calculation': f"Positive(${pos_total:,.2f}) + Negative(${neg_total:,.2f}) = ${net_amount:,.2f}",
                    'Status': 'Net Negative Balance'
                })
    
    return (pd.DataFrame(tgl_entries), 
            pd.DataFrame(matched_entries), 
            pd.DataFrame(unmatched_negatives))


def process_adjustments(combined_file: str, logger: logging.Logger) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Process adjustments data."""
    logger.info("Processing adjustments data...")
    
    try:
        # Read adjustments and tech data
        adj_df = pd.read_excel(combined_file, sheet_name='Direct Payroll Adjustments')
        tech_data = pd.read_excel(combined_file, sheet_name='Sheet1_Tech')
        
        # Filter out excluded techs
        tech_data = tech_data[~tech_data['Name'].isin(EXCLUDED_TECHS)]
        
        # Create tech lookup dictionary with proper columns
        tech_lookup = {}
        for _, row in tech_data.iterrows():
            badge_id = None
            if 'Badge ID' in tech_data.columns:
                badge_id = row['Badge ID'] if pd.notna(row['Badge ID']) else None
            if badge_id is None and 'Payroll ID' in tech_data.columns:
                badge_id = row['Payroll ID'] if pd.notna(row['Payroll ID']) else None
            
            # Get home department from Business Unit
            home_dept = get_tech_home_department(row['Technician Business Unit'])
            
            tech_lookup[row['Name']] = {
                'Badge ID': badge_id if badge_id is not None else '',
                'Technician Business Unit': row['Technician Business Unit'],
                'Home Department': home_dept
            }
        
        # Filter data
        adj_df = adj_df[
            (adj_df['Technician'].notna()) &
            (~adj_df['Technician'].astype(str).str.contains('Totals', na=False)) &
            (~adj_df['Technician'].isin(EXCLUDED_TECHS))
        ]
        
        # Process TGLs
        tgl_entries = []
        spiff_entries = []
        
        # First collect all negative spiffs by technician
        tech_negatives = defaultdict(float)
        
        for _, row in adj_df.iterrows():
            try:
                tech_name = row['Technician']
                tech_info = tech_lookup.get(tech_name)
                
                if not tech_info:
                    logger.debug(f"Skipping entry for tech not found in lookup: {tech_name}")
                    continue
                
                amount = float(str(row['Amount']).replace('$', '').replace(',', ''))
                memo = str(row['Memo']).strip()
                
                if 'tgl' in memo.lower():
                    subdept_code = memo[:2] if memo[:2].isdigit() else '00'
                    if subdept_code != '00':
                        tgl_entries.append({
                            'Technician': tech_name,
                            'Badge ID': tech_info['Badge ID'],
                            'Service Department': get_service_department_code(subdept_code),
                            'Amount': amount,
                            'Memo': memo,
                            'Type': 'TGL'
                        })
                else:
                    if amount < 0:
                        tech_negatives[tech_name] += amount
                    else:
                        subdept_code = memo[:2] if memo[:2].isdigit() else '00'
                        if subdept_code != '00':
                            spiff_entries.append({
                                'Technician': tech_name,
                                'Badge ID': tech_info['Badge ID'],
                                'Service Department': get_service_department_code(subdept_code),
                                'Amount': amount,
                                'Memo': memo,
                                'Type': 'Positive'
                            })
                
            except Exception as e:
                logger.warning(f"Error processing entry: {str(e)}")
                continue
        
        # Create negative entries
        negative_entries = []
        for tech_name, total_negative in tech_negatives.items():
            tech_info = tech_lookup[tech_name]
            negative_entries.append({
                'Technician': tech_name,
                'Badge ID': tech_info['Badge ID'],
                'Service Department': tech_info['Home Department'],
                'Amount': total_negative,
                'Type': 'Consolidated Negative',
                'Memo': f"Total negative spiffs to be subtracted from PCM"
            })
        
        # Convert to DataFrames
        tgl_df = pd.DataFrame(tgl_entries)
        spiff_df = pd.DataFrame(spiff_entries)
        neg_df = pd.DataFrame(negative_entries)
        
        logger.info(f"Successfully processed adjustments:")
        logger.info(f"  TGL entries: {len(tgl_df)}")
        logger.info(f"  Positive spiffs: {len(spiff_df)}")
        logger.info(f"  Techs with negative spiffs: {len(neg_df)}")
        
        return tgl_df, spiff_df, spiff_df, neg_df
        
    except Exception as e:
        logger.error(f"Error processing adjustments: {str(e)}")
        raise

def save_payroll_file(entries: List[PayrollEntry], output_file: str, logger: logging.Logger):
    """Save payroll entries to Excel file with specific formatting and validation."""
    try:
        # Define standard column order
        PAYROLL_COLUMNS = [
            'Company Code', 'Badge ID', 'Date', 'Amount', 
            'Pay Code', 'Dept', 'Location ID'
        ]
        
        # Convert entries to DataFrame
        df = pd.DataFrame([{
            'Company Code': entry.company_code,
            'Badge ID': entry.badge_id,
            'Date': entry.date,
            'Amount': entry.amount,
            'Pay Code': entry.pay_code,
            'Dept': entry.dept,
            'Location ID': entry.location_id
        } for entry in entries])
        
        # Group entries by Badge ID, Department, and Pay Code
        duplicate_check = df.groupby(['Badge ID', 'Dept', 'Pay Code', 'Company Code', 'Date', 'Location ID'])['Amount'].sum().reset_index()
        if len(duplicate_check) != len(df):
            logger.warning("Found multiple entries for same Badge ID, Department, and Pay Code. Consolidating...")
            df = duplicate_check
        
        # Sort entries and ensure column order
        df = df[PAYROLL_COLUMNS].sort_values(['Badge ID', 'Pay Code'])
        
        # Validate entries
        for idx, row in df.iterrows():
            try:
                # Ensure amount is positive
                if row['Amount'] <= 0:
                    logger.warning(f"Invalid amount ${row['Amount']} for Badge ID {row['Badge ID']}")
                    continue
                    
                # Validate Pay Code
                if row['Pay Code'] not in ['PCM', 'ICM', 'SPF']:
                    logger.warning(f"Invalid Pay Code {row['Pay Code']} for Badge ID {row['Badge ID']}")
                    continue
                    
                # Validate Department code
                if str(row['Dept']) not in [dept['code'] for dept in DEPARTMENT_CODES.values()]:
                    logger.warning(f"Invalid Department code {row['Dept']} for Badge ID {row['Badge ID']}")
                    continue
                
            except Exception as e:
                logger.error(f"Error validating row {idx}: {str(e)}")
                continue
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Format columns
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = max_length
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.font = Font(bold=True)
                
                # Special formatting for Amount column
                if col == 'Amount':
                    for cell in worksheet[col_letter]:
                        cell.alignment = Alignment(horizontal='right')
                        if cell.row > 1:  # Skip header
                            cell.number_format = '#,##0.00'
                
                # Center align other columns
                else:
                    for cell in worksheet[col_letter]:
                        cell.alignment = Alignment(horizontal='center')
        
        logger.info(f"Successfully saved {len(df)} payroll entries to {output_file}")
        logger.debug("Entry breakdown:")
        logger.debug(f"PCM (Service Tech Commission): {len(df[df['Pay Code'] == 'PCM'])}")
        logger.debug(f"ICM (Installer GP): {len(df[df['Pay Code'] == 'ICM'])}")
        logger.debug(f"SPF (Spiffs): {len(df[df['Pay Code'] == 'SPF'])}")
        
    except Exception as e:
        logger.error(f"Error saving payroll file: {str(e)}")
        raise

def save_adjustment_files(tgl_df: pd.DataFrame, matched_df: pd.DataFrame, 
                        pos_df: pd.DataFrame, neg_df: pd.DataFrame,
                        matched_file: str, pos_file: str, 
                        neg_file: str, tech_data: pd.DataFrame,
                        base_date: datetime, logger: logging.Logger):
    """Save adjustment files."""
    try:
        # Calculate week end date for entries
        start_of_week = base_date - timedelta(days=base_date.weekday())
        week_end_date = start_of_week + timedelta(days=6)
        target_date = week_end_date.strftime('%m/%d/%Y')
        
        # Initialize lists for tracking entries
        payroll_entries = []
        final_negative_entries = []
        
        # Read PCM entries from payroll file first
        payroll_file = os.path.join(os.path.dirname(matched_file), 'payroll.xlsx')
        try:
            all_payroll_entries = pd.read_excel(payroll_file, dtype={'Badge ID': str})
            # Filter for PCM entries only for processing negatives
            pcm_df = all_payroll_entries[all_payroll_entries['Pay Code'] == 'PCM'].copy()
            # Ensure Badge ID is properly formatted with THREE leading zeros
            pcm_df['Badge ID'] = pcm_df['Badge ID'].apply(format_badge_id)
        except Exception as e:
            logger.error(f"Error reading payroll file: {str(e)}")
            all_payroll_entries = pd.DataFrame()
            pcm_df = pd.DataFrame()
        
        # Helper function to normalize department codes (for comparison only)
        def normalize_dept(dept):
            try:
                # Convert to string, remove any leading/trailing spaces and leading zeros
                return str(int(str(dept).strip()))
            except (ValueError, TypeError):
                return str(dept).strip()
        
        # Process each tech's total negative against their PCM entry first
        updated_pcm_entries = []
        
        for _, neg_row in neg_df.iterrows():
            try:
                tech_badge_id = format_badge_id(neg_row['Badge ID'])
                home_dept = neg_row['Service Department']
                total_negative = abs(float(str(neg_row['Amount']).replace('$', '').replace(',', '')))
                
                # Look for PCM entry with matching badge ID and department
                if not pcm_df.empty:
                    # Create temporary normalized columns for comparison
                    normalized_home_dept = normalize_dept(home_dept)
                    temp_normalized = pcm_df['Dept'].apply(normalize_dept)
                    
                    # Debug logging
                    logger.debug(f"\nProcessing Badge ID: {tech_badge_id}")
                    logger.debug(f"Looking for department: {normalized_home_dept}")
                    logger.debug(f"Available PCM departments: {pcm_df['Dept'].tolist()}")
                    logger.debug(f"Normalized PCM departments: {temp_normalized.tolist()}")
                    
                    # Find matching entries using normalized departments
                    pcm_entries = pcm_df[
                        (pcm_df['Badge ID'] == tech_badge_id) &
                        (temp_normalized == normalized_home_dept)
                    ]
                    
                    logger.debug(f"Found matching PCM entries: {len(pcm_entries)}")
                    if not pcm_entries.empty:
                        logger.debug(f"PCM entries found:\n{pcm_entries}")
                    
                    if not pcm_entries.empty:
                        # Get the PCM amount
                        current_amount = float(str(pcm_entries.iloc[0]['Amount']).replace('$', '').replace(',', ''))
                        new_amount = current_amount - total_negative
                        
                        logger.debug(f"Current PCM amount: ${current_amount:,.2f}")
                        logger.debug(f"Negative amount to subtract: ${total_negative:,.2f}")
                        logger.debug(f"New amount after subtraction: ${new_amount:,.2f}")
                        
                        if new_amount > 0:
                            # Create updated PCM entry
                            updated_entry = pcm_entries.iloc[0].copy()
                            updated_entry['Amount'] = new_amount
                            updated_pcm_entries.append(updated_entry)
                            logger.debug(f"Updated PCM entry for Badge ID {tech_badge_id}: ${current_amount:,.2f} - ${total_negative:,.2f} = ${new_amount:,.2f}")
                            
                            # Remove this entry from pcm_df to prevent double processing
                            pcm_df = pcm_df.drop(pcm_entries.index)
                            continue  # Skip adding to final_negative_entries since we handled it
                    
                    # If we get here, either no PCM entry was found or new_amount <= 0
                    final_negative_entries.append({
                        'Technician': neg_row['Technician'],
                        'Badge ID': tech_badge_id,  # Already formatted with THREE leading zeros
                        'Service Department': home_dept,
                        'Amount': -total_negative,
                        'Type': 'Consolidated Negative',
                        'Memo': f"No matching PCM entry found for total negative spiffs of ${total_negative:,.2f}"
                    })
                
            except Exception as e:
                logger.warning(f"Error processing negative adjustment: {str(e)}")
                continue
        
        # Consolidate spiffs by Badge ID and Department
        spiff_groups = {}  # (badge_id, dept) -> total_amount
        
        # Process TGLs
        for _, row in tgl_df.iterrows():
            if pd.isna(row['Amount']):
                continue
                
            tech_name = row['Technician']
            if tech_name in EXCLUDED_TECHS:
                continue
            
            try:
                amount = float(str(row['Amount']).replace('$', '').replace(',', ''))
                badge_id = format_badge_id(row['Badge ID'])
                dept = row['Service Department']
                
                key = (badge_id, dept)
                spiff_groups[key] = spiff_groups.get(key, 0) + amount
                
            except (ValueError, TypeError) as e:
                logger.warning(f"Error processing TGL amount for {tech_name}: {str(e)}")
                continue
        
        # Process positive spiffs
        for _, row in matched_df.iterrows():
            if pd.isna(row['Amount']) or row['Amount'] <= 0:
                continue
                
            tech_name = row['Technician']
            if tech_name in EXCLUDED_TECHS:
                continue
                
            try:
                amount = float(str(row['Amount']).replace('$', '').replace(',', ''))
                badge_id = format_badge_id(row['Badge ID'])
                dept = row['Service Department']
                
                key = (badge_id, dept)
                spiff_groups[key] = spiff_groups.get(key, 0) + amount
                
            except (ValueError, TypeError) as e:
                logger.warning(f"Error processing positive spiff for {tech_name}: {str(e)}")
                continue
        
        # Convert consolidated spiffs to payroll entries
        for (badge_id, dept), amount in spiff_groups.items():
            payroll_entries.append({
                'Company Code': COMPANY_CODE,
                'Badge ID': badge_id,  # Already formatted
                'Date': target_date,
                'Amount': amount,
                'Pay Code': 'SPF',
                'Dept': dept,
                'Location ID': LOCATION_ID
            })
        
        # Update the payroll file with modified PCM entries
        if updated_pcm_entries:
            updated_pcm_df = pd.DataFrame(updated_pcm_entries)
            
            # Remove old PCM entries that were updated
            for _, updated_entry in updated_pcm_df.iterrows():
                mask = (
                    (all_payroll_entries['Badge ID'] == updated_entry['Badge ID']) &
                    (all_payroll_entries['Pay Code'] == 'PCM') &
                    (all_payroll_entries['Dept'] == updated_entry['Dept'])
                )
                all_payroll_entries = all_payroll_entries[~mask]
            
            # Add updated PCM entries
            all_payroll_entries = pd.concat([all_payroll_entries, updated_pcm_df], ignore_index=True)
            
            # Save back to payroll file
            all_payroll_entries = all_payroll_entries.sort_values(['Badge ID', 'Pay Code'])
            with pd.ExcelWriter(payroll_file, engine='openpyxl') as writer:
                all_payroll_entries.to_excel(writer, index=False)
                autofit_columns(writer.sheets['Sheet1'])
            
            logger.info(f"Updated {len(updated_pcm_entries)} PCM entries in payroll file")
        
        # Convert payroll entries to DataFrame
        payroll_df = pd.DataFrame(payroll_entries)
        
        # Save spiffs file
        if not payroll_df.empty:
            payroll_df = payroll_df.sort_values(['Badge ID', 'Dept'])
            
            with pd.ExcelWriter(matched_file, engine='openpyxl') as writer:
                payroll_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                
                for idx, col in enumerate(payroll_df.columns):
                    max_length = max(
                        payroll_df[col].astype(str).apply(len).max(),
                        len(str(col))
                    ) + 2
                    
                    col_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[col_letter].width = max_length
                    
                    header_cell = worksheet[f"{col_letter}1"]
                    header_cell.font = Font(bold=True)
                    
                    if col == 'Amount':
                        for cell in worksheet[col_letter]:
                            cell.alignment = Alignment(horizontal='right')
                            if cell.row > 1:
                                cell.number_format = '#,##0.00'
                    else:
                        for cell in worksheet[col_letter]:
                            cell.alignment = Alignment(horizontal='center')
        else:
            empty_df = pd.DataFrame(columns=[
                'Company Code', 'Badge ID', 'Date', 'Amount', 
                'Pay Code', 'Dept', 'Location ID'
            ])
            with pd.ExcelWriter(matched_file, engine='openpyxl') as writer:
                empty_df.to_excel(writer, index=False)
                autofit_columns(writer.sheets['Sheet1'])
        
        # Save reference files
        pos_reference_df = matched_df.copy()
        pos_reference_df['Processing Date'] = target_date
        
        with pd.ExcelWriter(pos_file, engine='openpyxl') as writer:
            pos_reference_df.to_excel(writer, index=False)
            autofit_columns(writer.sheets['Sheet1'])
        
        # Save negative adjustments
        neg_reference_df = pd.DataFrame(final_negative_entries)
        if not neg_reference_df.empty:
            neg_reference_df['Processing Date'] = target_date
        
        with pd.ExcelWriter(neg_file, engine='openpyxl') as writer:
            neg_reference_df.to_excel(writer, index=False)
            autofit_columns(writer.sheets['Sheet1'])
        
        logger.info(f"Successfully saved adjustment files:")
        logger.info(f"  Payroll entries: {len(payroll_df) if not payroll_df.empty else 0}")
        logger.info(f"  Positive reference entries: {len(pos_reference_df)}")
        logger.info(f"  Negative reference entries: {len(neg_reference_df)}")
        
    except Exception as e:
        logger.error(f"Error saving adjustment files: {str(e)}")
        raise

def validate_required_files(directory):
    """Validate that all required files are present before processing."""
    try:
        # Check for UUID file
        if not glob.glob(os.path.join(directory, "????????-????-????-????-????????????.xlsx")):
            raise FileNotFoundError("UUID file not found")
            
        # Check for Jobs Report
        if not glob.glob(os.path.join(directory, "Copy of Jobs Report for Performance -DE2_Dated *.xlsx")):
            raise FileNotFoundError("Jobs Report file not found")
            
        # Check for Tech Department file
        if not glob.glob(os.path.join(directory, "Technician Department_Dated *.xlsx")):
            raise FileNotFoundError("Tech Department file not found")
            
        # Check for Time Off file
        time_off_file = os.path.join(directory, "Approved_Time_Off 2023.xlsx")
        if not os.path.exists(time_off_file):
            raise FileNotFoundError("Time Off file not found")
            
        return True
        
    except FileNotFoundError as e:
        print(f"\nError: {str(e)}")
        print("Please ensure all required files are in the directory before running.")
        sys.exit(1)

def find_latest_files(directory):
    """Find the latest version of each required file."""
    try:
        uuid_file = glob.glob(os.path.join(directory, "????????-????-????-????-????????????.xlsx"))[0]
    except IndexError:
        raise FileNotFoundError("UUID file not found")
    
    try:
        jobs_file = max(glob.glob(os.path.join(directory, "Copy of Jobs Report for Performance -DE2_Dated *.xlsx")))
    except ValueError:
        raise FileNotFoundError("Jobs Report file not found")
    
    try:
        tech_file = max(glob.glob(os.path.join(directory, "Technician Department_Dated *.xlsx")))
    except ValueError:
        raise FileNotFoundError("Tech Department file not found")
    
    time_off_file = os.path.join(directory, "Approved_Time_Off 2023.xlsx")
    if not os.path.exists(time_off_file):
        raise FileNotFoundError("Time Off file not found")
    
    tgl_files = glob.glob(os.path.join(directory, "TGLs Set _Dated *.xlsx"))
    tgl_file = max(tgl_files) if tgl_files else None
    
    return {
        'uuid': uuid_file,
        'jobs': jobs_file,
        'tech': tech_file,
        'time_off': time_off_file,
        'tgl': tgl_file
    }

NAME_COLUMNS_MAP = {
    '2024': ['Technician Name'],
    'Sheet1_TGL': ['Lead Generated By'],
    'Invoices': ['Technician'],
    'Direct Payroll Adjustments': ['Technician'],
    'Sheet1': ['Sold By', 'Primary Technician'],
    'Sheet1_Tech': ['Name']
}

def combine_workbooks(directory, output_file, files):
    """Combine all workbooks into a single file using the pre-selected files."""
    def clean_name_column(worksheet, col_idx):
        """Helper function to clean names in a specific column."""
        for row in worksheet.iter_rows(min_row=2):  # Skip header
            cell = row[col_idx-1]  # Convert to 0-based index
            if cell.value and isinstance(cell.value, str):
                original = cell.value
                cleaned = original.strip()
                if original != cleaned:
                    print(f"Cleaned name: '{original}' -> '{cleaned}'")
                cell.value = cleaned

    # Copy the UUID file as the base
    shutil.copy2(files['uuid'], output_file)
    target_wb = load_workbook(filename=output_file)

    # Process the UUID file sheets first
    for sheet_name in target_wb.sheetnames:
        sheet = target_wb[sheet_name]
        header_row = next(sheet.rows)
        
        # Find Technician columns
        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'Technician':
                clean_name_column(sheet, idx)
        autofit_columns(sheet)

    # Process other workbooks
    source_configs = [
        {
            'file': files['jobs'],
            'source_sheet': 'Sheet1',
            'target_sheet': 'Sheet1',
            'name_cols': ['Sold By', 'Primary Technician', 'Technician']
        },
        {
            'file': files['tech'],
            'source_sheet': 'Sheet1',
            'target_sheet': 'Sheet1_Tech',
            'name_cols': ['Name']
        }
    ]

    if files.get('tgl'):
        source_configs.append({
            'file': files['tgl'],
            'source_sheet': 'Sheet1',
            'target_sheet': 'Sheet1_TGL',
            'name_cols': ['Lead Generated By']
        })

    for config in source_configs:
        source_wb = load_workbook(filename=config['file'], data_only=True)
        source_ws = source_wb[config['source_sheet']]

        # Create or replace target sheet
        if config['target_sheet'] in target_wb.sheetnames:
            target_wb.remove(target_wb[config['target_sheet']])
        target_ws = target_wb.create_sheet(config['target_sheet'])

        # Find columns that need cleaning
        name_col_indices = []
        header_row = next(source_ws.rows)
        for idx, cell in enumerate(header_row, 1):
            if cell.value in config['name_cols']:
                name_col_indices.append(idx)

        # Copy data and clean name columns
        for row in source_ws.rows:
            row_data = []
            for cell in row:
                value = cell.value
                if cell.row > 1 and cell.column in name_col_indices and isinstance(value, str):
                    original = value
                    value = value.strip()
                    if original != value:
                        print(f"Cleaned name: '{original}' -> '{value}'")
                target_ws.cell(row=cell.row, column=cell.column, value=value)

        autofit_columns(target_ws)

    target_wb.save(output_file)

def process_department_entries(tech_group: pd.DataFrame) -> Dict[str, Dict[str, float]]:
    """Process all entries for a technician by department, properly summing all positives and negatives."""
    dept_totals = defaultdict(lambda: {'positives': 0.0, 'negatives': 0.0})
    
    for _, row in tech_group.iterrows():
        try:
            amount = float(str(row['Amount']).replace('$', '').replace(',', ''))
            memo = str(row['Memo']).strip()
            
            # Skip TGLs
            if 'tgl' in memo.lower():
                continue
                
            # Extract department code
            dept_code = memo[:2]
            if not dept_code.isdigit():
                continue
                
            # Accumulate amounts by department
            if amount >= 0:
                dept_totals[dept_code]['positives'] += amount
            else:
                dept_totals[dept_code]['negatives'] += amount  # amount is already negative
                
        except (ValueError, AttributeError):
            continue
            
    return dept_totals

def process_calculations(base_path: str, output_dir: str, logger: logging.Logger,
                         start_of_week: datetime, end_of_week: datetime):
    """Process all calculations and generate output files."""
    try:
        combined_file = os.path.join(output_dir, 'combined_data.xlsx')
        paystats_file = os.path.join(output_dir, 'paystats.xlsx')

        # Read necessary data
        data = pd.read_excel(combined_file, sheet_name='Sheet1')
        tech_data = read_tech_department_data(combined_file, logger)

        # Filter tech_data to only include service technicians
        service_tech_data = tech_data[
            tech_data['Technician Business Unit'].apply(
                lambda x: determine_tech_type(x) == 'SERVICE'
            )
        ]

        # Use original time off file instead of combined file
        time_off_file = os.path.join(base_path, "Approved_Time_Off 2023.xlsx")
        excused_hours_dict = get_excused_hours(time_off_file, start_of_week)

        # Calculate commission results for service technicians only
        results_df = process_commission_calculations(
            data, service_tech_data, combined_file, start_of_week, excused_hours_dict
        )

        # Save results to paystats file
        with pd.ExcelWriter(paystats_file, engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Technician Revenue Totals', index=False)
            autofit_columns(writer.sheets['Technician Revenue Totals'])

        logger.info("Commission calculations completed for service technicians")

    except Exception as e:
        logger.error(f"Error in calculations: {str(e)}")
        raise

def process_payroll(base_path: str, output_dir: str, base_date: datetime, logger: logging.Logger, tech_data: pd.DataFrame):
    """Process payroll and adjustments, separating service tech and installer processing."""
    try:
        # Define file paths
        combined_file = os.path.join(output_dir, 'combined_data.xlsx')
        paystats_file = os.path.join(output_dir, 'paystats.xlsx')
        payroll_file = os.path.join(output_dir, 'payroll.xlsx')
        matched_file = os.path.join(output_dir, 'Spiffs.xlsx')
        adj_pos_file = os.path.join(output_dir, 'positive_adjustments.xlsx')
        adj_neg_file = os.path.join(output_dir, 'negative_adjustments.xlsx')
        
        # Format target date
        target_date = base_date.strftime('%m/%d/%Y')
        
        # Split technicians by type
        service_techs = tech_data[tech_data['Technician Business Unit'].apply(
            lambda x: determine_tech_type(x) == 'SERVICE'
        )]
        install_techs = tech_data[tech_data['Technician Business Unit'].apply(
            lambda x: determine_tech_type(x) == 'INSTALL'
        )]
        
        logger.info(f"Processing {len(service_techs)} service technicians and {len(install_techs)} installers")
        
        # Process service technician commissions
        payroll_entries = process_paystats(output_dir, paystats_file, service_techs, target_date, logger)
        
        # Process installer GP entries separately
        gp_entries = process_gp_entries(output_dir, install_techs, target_date, logger)
        
        # Combine payroll entries
        all_payroll_entries = payroll_entries + gp_entries
        
        # Process adjustments (TGLs and spiffs) for both service techs and installers
        eligible_techs = pd.concat([service_techs, install_techs])
        tgl_df, matched_df, pos_df, neg_df = process_adjustments(combined_file, logger, eligible_techs)
        
        # Save output files
        save_payroll_file(all_payroll_entries, payroll_file, logger)
        save_adjustment_files(tgl_df, matched_df, pos_df, neg_df, matched_file, 
                            adj_pos_file, adj_neg_file, tech_data, base_date, logger)
        
        logger.info("Payroll processing completed successfully!")
        logger.info(f"Service tech commission entries: {len(payroll_entries)}")
        logger.info(f"Installer GP entries: {len(gp_entries)}")
        logger.info(f"Total payroll entries: {len(all_payroll_entries)}")
        
    except Exception as e:
        logger.error(f"Error in payroll processing: {str(e)}")
        raise

def main():
    """Main program entry point with separated installer and service tech processing."""
    try:
        # Display welcome message and instructions
        print("\n" + "="*80)
        print("\nWelcome to the Commission Processing System\n")
        print("Before proceeding, please ensure the following files are in your Downloads folder:")
        print("\n1. UUID file (format: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx.xlsx)")
        print("2. Jobs Report (format: Copy of Jobs Report for Performance -DE2_Dated MM_DD_YY - MM_DD_YY.xlsx)")
        print("3. Tech Department file (format: Technician Department_Dated MM_DD_YY - MM_DD_YY.xlsx)")
        print("4. Time Off file (name: Approved_Time_Off 2023.xlsx)")
        print("5. TGL file (format: TGLs Set _Dated MM_DD_YY - MM_DD_YY.xlsx)")
        
        while True:
            response = input("\nAre all required files ready in the Downloads folder? (Y/N): ").strip().upper()
            if response == 'Y':
                break
            elif response == 'N':
                print("\nPlease gather all required files and run the program again.")
                sys.exit(0)
            else:
                print("Invalid input. Please enter Y or N.")

        base_path = os.path.expanduser(r'~\Downloads')
        
        # Get validated user date and files
        base_date, selected_uuid, found_files = get_validated_user_date(base_path)
        
        # Calculate week range
        start_of_week = base_date - timedelta(days=base_date.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        logger = setup_logging('commission_processor')
        logger.info("Starting commission processing...")
        logger.info(f"Using week range: {start_of_week.strftime('%m/%d/%Y')} to {end_of_week.strftime('%m/%d/%Y')}")

        output_dir = create_output_directory(base_path, start_of_week, end_of_week, logger)
        combined_file = os.path.join(output_dir, 'combined_data.xlsx')

        # Combine workbooks
        logger.info("Combining workbooks...")
        combine_workbooks(base_path, combined_file, found_files)
        logger.info("Workbook combination completed!")

        # Read and categorize technicians
        tech_data = read_tech_department_data(combined_file, logger)
        service_techs = tech_data[tech_data['Technician Business Unit'].apply(
            lambda x: determine_tech_type(x) == 'SERVICE'
        )]
        install_techs = tech_data[tech_data['Technician Business Unit'].apply(
            lambda x: determine_tech_type(x) == 'INSTALL'
        )]
        
        logger.info(f"\nFound {len(service_techs)} service technicians and {len(install_techs)} installers")
        
        # Process service technician calculations and paystats
        logger.info("\nProcessing service technician calculations...")
        process_calculations(base_path, output_dir, logger, start_of_week, end_of_week)

        # Process payroll entries for service technicians
        logger.info("\nProcessing service technician commission entries...")
        payroll_entries = process_paystats(
            output_dir, 
            os.path.join(output_dir, 'paystats.xlsx'), 
            tech_data,
            base_date,  # Pass the base_date
            logger
        )
        
        # Process installer GP entries
        logger.info("Processing installer GP entries...")
        gp_entries = process_gp_entries(
            output_dir, 
            install_techs, 
            base_date,  # Pass the base_date
            logger
        )
        
        # Process TGLs and spiffs for all eligible technicians
        logger.info("Processing TGLs and spiffs for all eligible technicians...")
        tgl_df, matched_df, pos_df, neg_df = process_adjustments(combined_file, logger)
        
        # Save final outputs
        all_payroll_entries = payroll_entries + gp_entries
        save_payroll_file(all_payroll_entries, os.path.join(output_dir, 'payroll.xlsx'), logger)
        
        # Save adjustment files with the base_date parameter
        save_adjustment_files(
            tgl_df, matched_df, pos_df, neg_df,
            os.path.join(output_dir, 'Spiffs.xlsx'),
            os.path.join(output_dir, 'positive_adjustments.xlsx'),
            os.path.join(output_dir, 'negative_adjustments.xlsx'),
            tech_data,
            base_date,  # Pass the base_date
            logger
        )

        logger.info("\nAll processing completed successfully!")
        logger.info(f"Service tech commission entries: {len(payroll_entries)}")
        logger.info(f"Installer GP entries: {len(gp_entries)}")
        logger.info(f"Total payroll entries: {len(all_payroll_entries)}")

    except Exception as e:
        logger.error(f"Fatal error in main process: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()