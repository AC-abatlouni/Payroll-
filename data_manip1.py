import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
import sys
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, List

COLUMN_ORDER = [
    'Badge ID', 'Technician', 'Main Dept',
    # HVAC Department Subdepartments
    '20 Revenue', '20 Sales', '20 Total',
    '21 Revenue', '21 Sales', '21 Total',
    '22 Revenue', '22 Sales', '22 Total',
    '24 Revenue', '24 Sales', '24 Total',
    '25 Revenue', '25 Sales', '25 Total',
    '27 Revenue', '27 Sales', '27 Total',
    # HVAC Department Totals
    'HVAC Revenue', 'HVAC Sales', 'HVAC Spiffs', 'HVAC Total', 'HVAC Commission',
    # Plumbing Department Subdepartments
    '30 Revenue', '30 Sales', '30 Total',
    '31 Revenue', '31 Sales', '31 Total',
    '33 Revenue', '33 Sales', '33 Total',
    '34 Revenue', '34 Sales', '34 Total',  # Added new excavation department
    # Plumbing Department Totals
    'Plumbing Revenue', 'Plumbing Sales', 'Plumbing Spiffs', 'Plumbing Total', 'Plumbing Commission',
    # Electric Department Subdepartments
    '40 Revenue', '40 Sales', '40 Total',
    '41 Revenue', '41 Sales', '41 Total',
    # Electric Department Totals
    'Electric Revenue', 'Electric Sales', 'Electric Spiffs', 'Electric Total', 'Electric Commission',
    # Performance Metrics
    'Completed Job Revenue', 'Tech-Sourced Install Sales',
    'Service Completion %', 'Install Contribution %',
    'Excused Hours', 'Spiffs', 'Valid TGLs', 'Avg Ticket $',
    'TGL Threshold Reduction', 'Base Threshold Scale', 'Adjusted Threshold Scale',
    'Commission Rate %', 'Total Revenue', 'Commissionable Revenue',
    'Status', 'Total Commission'
]

# Define threshold tables
hvac_thresholds = {
    0: [7000, 8000, 9000, 10000],
    10: [7500, 8500, 9500, 11000],
    20: [8000, 9000, 10000, 11000],
    30: [9000, 10000, 11000, 12000],
    40: [10000, 11000, 12000, 13000],
    50: [12000, 14000, 16000, 18000],
    60: [14000, 16000, 18000, 20000],
    70: [15000, 17000, 19000, 21000],
    80: [16500, 18500, 20500, 23000],
    90: [18500, 20500, 23500, 26000],
    100: [22000, 24000, 26000, 29000]
}

plumbing_electrical_thresholds = {
    0: [7000, 8000, 9000, 10000],
    10: [7500, 8500, 9500, 11000],
    20: [8000, 9000, 10000, 11000],
    30: [9000, 10000, 11000, 12000],
    40: [10000, 11000, 12000, 13000],
    50: [11500, 12500, 14000, 15000],
    60: [13000, 14000, 15500, 17000],
    70: [14500, 15500, 17000, 18000],
    80: [15500, 16500, 18000, 19000],
    90: [17000, 18500, 19500, 21000],
    100: [18000, 20000, 22000, 24000]
}


SUBDEPARTMENT_MAP = {
    '00': '00 - ADMINISTRATIVE',
    '41': '41 - ELECTRICAL INSTALL',
    '40': '40 - ELECTRICAL SERVICE',
    '42': '42 - GENERATOR MAINTENANCE',
    '27': '27 - HVAC DUCT CLEANING',
    '21': '21 - HVAC INSTALL',
    '23': '23 - HVAC SALES',
    '20': '20 - HVAC SERVICE',
    '22': '22 - MAINTENANCE MVP',
    '25': '25 - OIL MAINTENANCE MVP',
    '24': '24 - OIL SERVICE',
    '33': '33 - PLUMBING DRAIN CLEANING',
    '34': '34 - PLUMBING EXCAVATION',  
    '31': '31 - PLUMBING INSTALL',
    '32': '32 - PLUMBING MAINTENANCE',
    '30': '30 - PLUMBING SERVICE'
}

def get_tech_data(file_path: str) -> pd.DataFrame:
    """Read technician data from Sheet1_Tech."""
    try:
        tech_data = pd.read_excel(file_path, sheet_name='Sheet1_Tech')
        required_columns = ['Name', 'Payroll ID', 'Technician Business Unit']
        missing_columns = [col for col in required_columns if col not in tech_data.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns in Sheet1_Tech: {missing_columns}")
        return tech_data
    except Exception as e:
        logger.error(f"Error reading technician data: {str(e)}")
        raise

def setup_logging():
    logger = logging.getLogger('commission_calculator')
    logger.setLevel(logging.DEBUG)
    
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    fh = logging.FileHandler(f'commission_calculations_{current_time}.log')
    fh.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    fh.setFormatter(file_formatter)
    
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(message)s')
    ch.setFormatter(console_formatter)
    
    class ConsoleFilter(logging.Filter):
        def filter(self, record):
            return record.levelno == logging.INFO and not any(x in record.msg for x in [
                'Department Summary',
                'Completed:',
                'Sales:',
                'Total:',
                'Spiffs:',
                'Commissionable Revenue:',
                'Commission:',
                'Processing technician'
            ])
    
    ch.addFilter(ConsoleFilter())
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

logger = setup_logging()

def format_currency(amount):
    try:
        if isinstance(amount, str):
            amount = float(amount.replace('$', '').replace(',', ''))
        return f"${amount:,.2f}"
    except (ValueError, TypeError):
        return "$0.00"

def get_user_date() -> datetime:
    while True:
        try:
            date_str = input("Enter a date (mm/dd/yy): ")
            return datetime.strptime(date_str, '%m/%d/%y')
        except ValueError:
            print("Invalid date format. Please use mm/dd/yy format (e.g., 03/15/24)")

def extract_subdepartment_code(business_unit):
    """Extract specific two-digit subdepartment code from business unit."""
    try:
        unit_str = str(business_unit).split('-')[0].strip()
        subdept = ''.join(filter(str.isdigit, unit_str))
        return subdept[:2] if len(subdept) >= 2 else '00'
    except:
        return '00'

def extract_department_number(business_unit):
    try:
        unit_str = str(business_unit).split('-')[0].strip()
        return int(''.join(filter(str.isdigit, unit_str)))
    except:
        return 0

def get_department_with_code(dept_num):
    if 20 <= dept_num <= 29:
        return 'HVAC (20-29)'
    elif 30 <= dept_num <= 39:
        return 'Plumbing (30-39)'
    elif 40 <= dept_num <= 49:
        return 'Electric (40-49)'
    return 'Unknown (0)'

def get_department_from_number(dept_num):
    if 20 <= dept_num <= 29:
        return 'HVAC'
    elif 30 <= dept_num <= 39:
        return 'Plumbing'
    elif 40 <= dept_num <= 49:
        return 'Electric'
    return 'Unknown'

def autofit_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def extract_department_range(business_unit):
    try:
        unit_number = int(''.join(filter(str.isdigit, business_unit)))
        base = (unit_number // 10) * 10
        if 20 <= base <= 20:
            return range(20, 30)
        elif 30 <= base <= 30:
            return range(30, 40)
        elif 40 <= base <= 40:
            return range(40, 50)
        return range(0, 0)
    except (ValueError, TypeError):
        logger.warning(f"Could not extract department range from business unit: {business_unit}")
        return range(0, 0)

def is_same_department(source_unit, target_unit):
    source_range = extract_department_range(source_unit)
    target_range = extract_department_range(target_unit)
    
    return (len(source_range) > 0 and 
            len(target_range) > 0 and 
            source_range.start == target_range.start)

def format_threshold_scale(thresholds):
    return (f"2%: ${thresholds[0]:,.0f} | "
            f"3%: ${thresholds[1]:,.0f} | "
            f"4%: ${thresholds[2]:,.0f} | "
            f"5%: ${thresholds[3]:,.0f}")

def get_valid_tgls(file_path: str, tech_name: str) -> List[dict]:
    try:
        tgl_df = pd.read_excel(file_path, sheet_name='Sheet1_TGL')
        logger.debug(f"Processing TGLs for {tech_name} from Sheet1_TGL")
        
        valid_tgls = []
        tech_tgls = tgl_df[
            (tgl_df['Lead Generated By'] == tech_name) & 
            (tgl_df['Status'] != 'Canceled')
        ]
        
        for _, tgl in tech_tgls.iterrows():
            source_unit = str(tgl['Business Unit'])
            target_unit = str(tgl['Lead Generated from Business Unit'])
            
            if is_same_department(source_unit, target_unit):
                valid_tgls.append({
                    'job_number': tgl.get('Job #', 'N/A'),
                    'status': tgl['Status'],
                    'business_unit': source_unit,
                    'target_unit': target_unit,
                    'created_date': tgl['Created Date']
                })
                logger.debug(f"Valid TGL found - Job #: {tgl.get('Job #', 'N/A')}, "
                           f"From: {source_unit} To: {target_unit}")
        
        return valid_tgls
        
    except Exception as e:
        logger.error(f"Error processing TGL data for {tech_name}: {str(e)}")
        return []

def get_spiffs_total(file_path: str, tech_name: str) -> tuple[float, dict[str, float]]:
    try:
        spiffs_df = pd.read_excel(file_path, sheet_name='Direct Payroll Adjustments')
        tech_spiffs = spiffs_df[spiffs_df['Technician'] == tech_name]
        
        spiffs_total = 0
        department_spiffs = {
            'HVAC': 0,
            'Plumbing': 0,
            'Electric': 0
        }
        
        for idx, spiff in tech_spiffs.iterrows():
            amount = float(str(spiff['Amount']).replace('$', '').replace(',', '').strip()) if pd.notnull(spiff['Amount']) else 0
            memo = str(spiff['Memo']).strip() if pd.notnull(spiff['Memo']) else ''
            
            # Skip if memo contains 'commission' or 'tgl' in any case
            if 'commission' in memo.lower() or 'tgl' in memo.lower():
                logger.debug(f"Skipping entry: ${amount:,.2f} - Memo: {memo}")
                continue
            
            if not memo[:2].isdigit():
                logger.error(f"Invalid memo format - must start with department number. Row {idx + 2}: {memo}")
                raise ValueError(f"Invalid memo format - must start with department number. Row {idx + 2}: {memo}")
            
            dept_num = int(memo[:2])
            if not (20 <= dept_num <= 29 or 30 <= dept_num <= 39 or 40 <= dept_num <= 49):
                logger.error(f"Invalid department number in memo (must be 20-29, 30-39, or 40-49). Row {idx + 2}: {memo}")
                raise ValueError(f"Invalid department number in memo (must be 20-29, 30-39, or 40-49). Row {idx + 2}: {memo}")
            
            memo_content = memo[2:].strip().lstrip('-').strip()
            if not memo_content:
                logger.warning(f"Warning: Memo contains no description after department number. Row {idx + 2}: {memo}")
            
            if 20 <= dept_num <= 29:
                department_spiffs['HVAC'] += amount
                logger.debug(f"Added HVAC spiff: ${amount:,.2f} - Memo: {memo}")
            elif 30 <= dept_num <= 39:
                department_spiffs['Plumbing'] += amount
                logger.debug(f"Added Plumbing spiff: ${amount:,.2f} - Memo: {memo}")
            elif 40 <= dept_num <= 49:
                department_spiffs['Electric'] += amount
                logger.debug(f"Added Electric spiff: ${amount:,.2f} - Memo: {memo}")
            
            spiffs_total += amount
        
        return spiffs_total, department_spiffs
        
    except Exception as e:
        logger.error(f"Error processing spiffs data for {tech_name}: {str(e)}")
        raise

def get_excused_hours(file_path: str, base_date: datetime, sheet_name: str = '2024') -> Dict[str, int]:
    """Get excused hours from time off sheet."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        start_of_week = base_date - timedelta(days=base_date.weekday())
        end_of_week = start_of_week + timedelta(days=4)
        
        def format_date(date):
            day = date.day
            if 4 <= day <= 20 or 24 <= day <= 30:
                suffix = "th"
            else:
                suffix = ["st", "nd", "rd"][day % 10 - 1]
            return f"{date.strftime('%B')} {day}{suffix}"
        
        target_week = f"{format_date(start_of_week)} - {format_date(end_of_week)}"
        logger.debug(f"Looking for week range: {target_week}")
        
        start_col = None
        for row_idx in [0, 1]:
            for col_idx, header in enumerate(df.iloc[row_idx]):
                if isinstance(header, str) and target_week.strip() == header.strip():
                    start_col = col_idx
                    logger.debug(f"Found target week in row {row_idx}, column {col_idx}")
                    break
            if start_col is not None:
                break
        
        if start_col is None:
            logger.warning(f"Week range '{target_week}' not found in time off sheet")
            return {}
        
        week_columns = list(range(start_col, start_col + 5))
        hours_summary = {}
        
        for index in range(2, len(df)):
            row = df.iloc[index]
            technician_name = row[0]
            
            if pd.isna(technician_name) or str(technician_name).strip() == '':
                continue
            
            technician_name = str(technician_name).strip()
            total_hours = 0
            
            for col in week_columns:
                try:
                    cell_value = str(row[col]).strip().lower()
                    if cell_value in ['x', 'r', 'v']:
                        total_hours += 8
                        logger.debug(f"Found time off marker '{cell_value}' for {technician_name} in column {col}")
                except Exception as e:
                    logger.warning(f"Error processing cell for {technician_name} in column {col}: {str(e)}")
                    continue
            
            if total_hours > 0:
                hours_summary[technician_name] = total_hours
                logger.debug(f"Found {total_hours} excused hours for {technician_name}")
        
        return hours_summary
        
    except Exception as e:
        logger.error(f"Error analyzing time off data: {str(e)}")
        return {}

def calculate_box_metrics(data: pd.DataFrame, tech_name: str, base_date: datetime) -> Tuple[float, float, float, Dict[str, Dict[str, float]]]:
    """Calculate Box A (CJR), Box B (TSIS), and Box C (Total) metrics with subdepartment breakdowns."""
    subdept_breakdown = {
        'completed': {code: 0.0 for code in SUBDEPARTMENT_MAP.keys()},
        'sales': {code: 0.0 for code in SUBDEPARTMENT_MAP.keys()},
        'total': {code: 0.0 for code in SUBDEPARTMENT_MAP.keys()}
    }

    # Calculate week range
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    logger.debug(f"\nCalculating metrics for {tech_name} for week {start_of_week.strftime('%Y-%m-%d')} to {end_of_week.strftime('%Y-%m-%d')}")
    
    # Ensure datetime format
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])

    # Get primary tech jobs within date range
    primary_jobs = data[
        (data['Primary Technician'] == tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    box_a = primary_jobs['Jobs Total Revenue'].fillna(0).sum()
    
    # Get sales jobs within date range
    sold_jobs = data[
        (data['Sold By'] == tech_name) & 
        (data['Primary Technician'] != tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    box_b = sold_jobs['Jobs Total Revenue'].fillna(0).sum()
    
    # Calculate subdepartment breakdowns
    for _, row in primary_jobs.iterrows():
        revenue = row['Jobs Total Revenue'] or 0
        subdept = extract_subdepartment_code(row.get('Business Unit', ''))
        subdept_breakdown['completed'][subdept] += revenue
        subdept_breakdown['total'][subdept] += revenue
        
    for _, row in sold_jobs.iterrows():
        revenue = row['Jobs Total Revenue'] or 0
        subdept = extract_subdepartment_code(row.get('Business Unit', ''))
        subdept_breakdown['sales'][subdept] += revenue
        subdept_breakdown['total'][subdept] += revenue

    box_c = box_a + box_b
    
    return box_a, box_b, box_c, subdept_breakdown

def calculate_percentages(box_a: float, box_c: float) -> Tuple[int, int]:
    """Calculate Service Completion and Install Contribution percentages."""
    if box_c == 0:
        return 0, 0
        
    raw_scp = (box_a / box_c * 100)
    raw_icp = 100 - raw_scp
    
    raw_scp = max(0, raw_scp)
    raw_icp = max(0, raw_icp)
    
    total = raw_scp + raw_icp
    if total != 0:
        raw_scp = (raw_scp / total) * 100
        raw_icp = (raw_icp / total) * 100
        
    scp = round(raw_scp / 10) * 10
    icp = round(raw_icp / 10) * 10
    
    if scp + icp != 100:
        if scp > icp:
            scp = 100
            icp = 0
        else:
            scp = 0
            icp = 100
    
    return int(scp), int(icp)

def calculate_average_ticket_value(data: pd.DataFrame, tech_name: str, box_a: float, box_b: float, base_date: datetime, logger: logging.Logger) -> Dict[str, float]:
    """Calculate average ticket value using total revenue divided by opportunity count."""
    avg_tickets = {'overall': 0.0}
    
    # Calculate week range
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    logger.debug(f"\nCalculating average ticket value for {tech_name}")
    logger.debug(f"Week range: {start_of_week.date()} to {end_of_week.date()}")
    
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])

    # Get completed jobs within date range to count opportunities
    completed_jobs = data[
        (data['Primary Technician'] == tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    # Get count of opportunity jobs
    opportunity_jobs = completed_jobs[completed_jobs['Opportunity'] == True]
    opportunity_count = len(opportunity_jobs)
    
    # Calculate average ticket using total revenue
    total_revenue = box_a + box_b
    avg_ticket = round(total_revenue / opportunity_count, 2) if opportunity_count > 0 else 0
    avg_tickets['overall'] = avg_ticket
    
    logger.debug(f"Total Revenue (CJR + TSIS): ${total_revenue:,.2f}")
    logger.debug(f"True Opportunity Count: {opportunity_count}")
    logger.debug(f"Average Ticket: ${avg_ticket:,.2f}")
    
    # Log opportunity jobs for reference
    logger.debug("\nOpportunity Jobs:")
    for _, job in opportunity_jobs.iterrows():
        invoice = job.get('Invoice #', 'N/A')
        customer_name = job.get('Customer Name', 'Unknown')
        invoice_date = job['Invoice Date'].strftime('%m/%d/%y')
        logger.debug(f"  Invoice #{invoice} - {invoice_date} - {customer_name}")
    
    return avg_tickets
    
def format_department_revenue(revenue_data: Dict[str, Dict[str, float]], 
                            commission_rate: float,
                            department_spiffs: Dict[str, float],
                            subdept_breakdown: Dict[str, Dict[str, float]]) -> Dict[str, str]:
    """Format department and subdepartment revenue data into strings for Excel output."""
    formatted = {}
    
    # Format subdepartment totals
    for subdept_code in ['20', '21', '22', '24', '25', '27', '30', '31', '33', '34', '40', '41']:  # Added '34'
        completed = subdept_breakdown['completed'].get(subdept_code, 0)
        sales = subdept_breakdown['sales'].get(subdept_code, 0)
        total = subdept_breakdown['total'].get(subdept_code, 0)
        
        formatted[f"{subdept_code} Revenue"] = f"${completed:,.2f}"
        formatted[f"{subdept_code} Sales"] = f"${sales:,.2f}"
        formatted[f"{subdept_code} Total"] = f"${total:,.2f}"
    
    # Format main department totals
    for dept in ['HVAC', 'Plumbing', 'Electric']:
        completed = revenue_data['completed'][dept]
        sales = revenue_data['sales'][dept]
        combined = revenue_data['combined'][dept]
        dept_spiffs = department_spiffs[dept]
        adjusted_combined = combined - dept_spiffs
        
        formatted[f"{dept} Revenue"] = f"${completed:,.2f}"
        formatted[f"{dept} Sales"] = f"${sales:,.2f}"
        formatted[f"{dept} Spiffs"] = f"${dept_spiffs:,.2f}"
        formatted[f"{dept} Total"] = f"${adjusted_combined:,.2f}"
        formatted[f"{dept} Commission"] = f"${(adjusted_combined * commission_rate):,.2f}"
    
    return formatted

def calculate_department_revenue(data: pd.DataFrame, tech_name: str, base_date: datetime) -> Dict[str, Dict[str, float]]:
    """Calculate department revenue using the same date filtering."""
    # Calculate week range
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    
    revenue_by_dept = {
        'completed': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0},
        'sales': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0},
        'combined': {'HVAC': 0.0, 'Plumbing': 0.0, 'Electric': 0.0, 'Unknown': 0.0}
    }
    
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])

    # Get completed jobs within date range
    completed_jobs = data[
        (data['Primary Technician'] == tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    # Get sales within date range
    sold_jobs = data[
        (data['Sold By'] == tech_name) & 
        (data['Primary Technician'] != tech_name) &
        (data['Invoice Date'].dt.date >= start_of_week.date()) &
        (data['Invoice Date'].dt.date <= end_of_week.date())
    ]
    
    # Calculate revenues
    for _, job in completed_jobs.iterrows():
        dept_num = extract_department_number(str(job.get('Business Unit', '')))
        dept = get_department_from_number(dept_num)
        revenue = job.get('Jobs Total Revenue', 0) or 0
        revenue_by_dept['completed'][dept] += revenue
        revenue_by_dept['combined'][dept] += revenue

    for _, job in sold_jobs.iterrows():
        dept_num = extract_department_number(str(job.get('Business Unit', '')))
        dept = get_department_from_number(dept_num)
        revenue = job.get('Jobs Total Revenue', 0) or 0
        revenue_by_dept['sales'][dept] += revenue
        revenue_by_dept['combined'][dept] += revenue

    return revenue_by_dept

def get_commission_rate(total_revenue: float, flipped_percent: float, department: str, 
                       excused_hours: int, tgl_reduction: float) -> Tuple[float, list, list]:
    """Calculate commission rate and thresholds based on revenue and department."""
    logger.debug(f"\nDetailed threshold calculation:")
    
    flipped_percent = min(100, max(0, int(round(flipped_percent / 10) * 10)))
    
    if department in ['Electric', 'Plumbing']:
        thresholds = plumbing_electrical_thresholds
    else:
        thresholds = hvac_thresholds
    
    tier_thresholds = thresholds[flipped_percent].copy()
    logger.debug(f"Using {department} thresholds for ICP {flipped_percent}%: {tier_thresholds}")
    
    days_off = min(5, excused_hours / 8)
    reduction_factor = max(0, 1 - (0.20 * days_off))
    logger.debug(f"Time off reduction factor: {reduction_factor} ({days_off} days)")
    
    # First apply time off reduction
    time_off_adjusted = [threshold * reduction_factor for threshold in tier_thresholds]
    logger.debug(f"After time off adjustment: {time_off_adjusted}")
    
    # Then apply TGL reduction
    adjusted_thresholds = [max(0, threshold - tgl_reduction) for threshold in time_off_adjusted]
    logger.debug(f"After TGL reduction: {adjusted_thresholds}")
    
    # Determine commission rate based on highest threshold met
    if total_revenue >= adjusted_thresholds[3]:
        rate = 0.05
    elif total_revenue >= adjusted_thresholds[2]:
        rate = 0.04
    elif total_revenue >= adjusted_thresholds[1]:
        rate = 0.03
    elif total_revenue >= adjusted_thresholds[0]:
        rate = 0.02
    else:
        rate = 0
    
    logger.debug(f"Final commission rate: {rate}")
    return rate, adjusted_thresholds, tier_thresholds

def main():
    technicians_to_track = []  # Will be populated from Sheet1_Tech
    
    file_path = r'C:\Users\abatlouni\Downloads\combined_data.xlsx'
    logger.info(f"Loading data from {file_path}")
    
    # Get the base date once at the start
    base_date = get_user_date()
    
    try:
        data = pd.read_excel(file_path, sheet_name='Sheet1')
        tech_data = get_tech_data(file_path)
        
        # Get list of technicians from Sheet1_Tech
        technicians_to_track = tech_data['Name'].tolist()
        
    except Exception as e:
        logger.error(f"Error loading Excel sheets: {e}")
        sys.exit(1)
    
    # Ensure Invoice Date is in datetime format right at the start
    if not pd.api.types.is_datetime64_any_dtype(data['Invoice Date']):
        data['Invoice Date'] = pd.to_datetime(data['Invoice Date'])
    
    logger.debug(f"Columns in 'Sheet1_Tech': {tech_data.columns.tolist()}")
    
    # Pass base_date to get_excused_hours
    excused_hours_dict = get_excused_hours(file_path, base_date)
    
    tech_dept_map = dict(zip(tech_data['Name'], tech_data['Technician Business Unit']))
    tech_badge_map = dict(zip(tech_data['Name'], tech_data['Payroll ID']))
    
    results = []

    for tech_name in technicians_to_track:
        logger.info(f"\nProcessing technician: {tech_name}")
        badge_id = tech_badge_map.get(tech_name, '')  # Get badge ID from Sheet1_Tech data
        
        # Rest of the processing remains the same...
        box_a, box_b, box_c, subdept_breakdown = calculate_box_metrics(data, tech_name, base_date)
        scp, icp = calculate_percentages(box_a, box_c)
        
        dept_revenue = calculate_department_revenue(data, tech_name, base_date)
        spiffs_total, department_spiffs = get_spiffs_total(file_path, tech_name)
        valid_tgls = get_valid_tgls(file_path, tech_name)
        avg_tickets = calculate_average_ticket_value(data, tech_name, box_a, box_b, base_date, logger)
        default_ticket = 0.0
        avg_ticket_value = avg_tickets.get('overall', default_ticket) if avg_tickets else default_ticket
        
        dept_unit_str = tech_dept_map.get(tech_name, '0')
        dept_num = extract_department_number(dept_unit_str)
        department = get_department_from_number(dept_num)
        department_with_code = get_department_with_code(dept_num)

        tgl_reduction = avg_ticket_value * len(valid_tgls) if avg_ticket_value > 0 else 0
        excused_hours = excused_hours_dict.get(tech_name, 0)
        
        commission_rate, adjusted_thresholds, base_thresholds = get_commission_rate(
            box_c, icp, department, excused_hours, tgl_reduction
        )
        
        base_threshold_scale = format_threshold_scale(base_thresholds)
        adjusted_threshold_scale = format_threshold_scale(adjusted_thresholds)
        
        formatted_dept_data = format_department_revenue(
            dept_revenue,
            commission_rate,
            department_spiffs,
            subdept_breakdown
        )
        
        commissionable_revenue = box_c - spiffs_total
        final_commission = commissionable_revenue * commission_rate
        
        result = {
            'Badge ID': badge_id,
            'Technician': tech_name,
            'Main Dept': department_with_code,
            'Total Revenue': box_c,
            'Completed Job Revenue': box_a,
            'Tech-Sourced Install Sales': box_b,
            'Service Completion %': scp,
            'Install Contribution %': icp,
            'Excused Hours': excused_hours,
            'Spiffs': spiffs_total,
            'Valid TGLs': len(valid_tgls),
            'Avg Ticket $': avg_ticket_value,
            'TGL Threshold Reduction': tgl_reduction,
            'Base Threshold Scale': base_threshold_scale,
            'Adjusted Threshold Scale': adjusted_threshold_scale,
            'Commissionable Revenue': commissionable_revenue,
            'Commission Rate %': commission_rate * 100,
            'Total Commission': round(final_commission, 2),
            'Status': f"Qualified for {commission_rate*100}% tier" if commission_rate > 0 else "Did not qualify"
        }
        
        result.update(formatted_dept_data)
        results.append(result)

    results_df = pd.DataFrame(results)
    
    # Ensure all required columns exist (fill with empty values if missing)
    for col in COLUMN_ORDER:
        if col not in results_df.columns:
            results_df[col] = ''
    
    # Reorder columns
    results_df = results_df[COLUMN_ORDER]
    
    output_path = r'C:\Users\abatlouni\Downloads\paystats.xlsx'
    logger.info(f"\nWriting results to {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Technician Revenue Totals', index=False)
        wb = writer.book
        ws = wb['Technician Revenue Totals']
        autofit_columns(ws)

    logger.info("Commission calculations completed")

if __name__ == "__main__":
    main()