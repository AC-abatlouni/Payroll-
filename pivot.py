import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
import sys
from datetime import datetime, timedelta
from typing import Dict, Optional

# Set up logging
def setup_logging():
    logger = logging.getLogger('commission_calculator')
    logger.setLevel(logging.DEBUG)
    
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    fh = logging.FileHandler(f'commission_calculations_{current_time}.log')
    fh.setLevel(logging.DEBUG)
    
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_formatter = logging.Formatter('%(levelname)s - %(message)s')
    fh.setFormatter(file_formatter)
    ch.setFormatter(console_formatter)
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    
    return logger

def get_week_range(reference_date: Optional[datetime] = None) -> str:
    """Calculate the week range based on reference date."""
    base_date = reference_date or datetime.today()
    start_of_week = base_date - timedelta(days=base_date.weekday())
    end_of_week = start_of_week + timedelta(days=4)
    
    def format_date(date):
        day = date.day
        if 4 <= day <= 20 or 24 <= day <= 30:
            suffix = "th"
        else:
            suffix = ["st", "nd", "rd"][day % 10 - 1]
        return f"{date.strftime('%B')} {day}{suffix}"
    
    return f"{format_date(start_of_week)} - {format_date(end_of_week)}"

def get_excused_hours(file_path: str, sheet_name: str = '2024') -> Dict[str, int]:
    """
    Get excused hours for each technician for the current week.
    
    Args:
        file_path: Path to the Excel file containing time off data
        sheet_name: Sheet name to analyze (defaults to '2024')
    
    Returns:
        Dict[str, int]: Dictionary mapping technician names to their excused hours
    """
    try:
        # Load the time off data
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Get the target week range
        base_date = datetime.today()
        start_of_week = base_date - timedelta(days=base_date.weekday())
        end_of_week = start_of_week + timedelta(days=4)
        
        def format_date(date):
            day = date.day
            if 4 <= day <= 20 or 24 <= day <= 30:
                suffix = "th"
            else:
                suffix = ["st", "nd", "rd"][day % 10 - 1]
            return f"{date.strftime('%B')} {day}{suffix}"
        
        target_week = f"{format_date(start_of_week)} - {format_date(end_of_week)}"
        
        # Find the start column for the target week
        start_col = None
        for i, header in enumerate(df.iloc[0]):
            if isinstance(header, str) and target_week.strip() == header.strip():
                start_col = i
                break
        
        if start_col is None:
            logger.warning(f"Week range '{target_week}' not found in time off sheet")
            return {}
        
        # Get the weekday columns
        week_columns = df.columns[start_col:start_col + 5]
        
        # Calculate hours for each technician
        hours_summary = {}
        
        # Start from row 2 to skip headers
        for index, row in df.iloc[2:].iterrows():
            technician_name = row[0]
            
            if pd.isna(technician_name):
                continue
                
            technician_name = str(technician_name).strip()
            total_hours = 0
            
            # Check each day for time off markers
            for day in week_columns:
                cell_value = str(row[day]).strip().lower()
                if cell_value in ['x', 'r']:
                    total_hours += 8
            
            if total_hours > 0:
                hours_summary[technician_name] = total_hours
                logger.debug(f"Found {total_hours} excused hours for {technician_name}")
        
        logger.info(f"Found excused hours for {len(hours_summary)} technicians")
        return hours_summary
        
    except Exception as e:
        logger.error(f"Error analyzing time off data: {str(e)}")
        return {}

def custom_round_percentage(value):
    """Round percentage to nearest 10% using round half up method."""
    logger.debug(f"Rounding percentage value: {value}")
    if pd.isna(value):
        logger.debug("Found NaN value, returning 0")
        return 0
    value = max(0, min(100, value))
    logger.debug(f"Clamped value between 0-100: {value}")
    rounded_value = int(value // 10 * 10)
    if value % 10 >= 5:
        rounded_value += 10
    logger.debug(f"Final rounded value: {rounded_value}")
    return rounded_value

def get_commission_rate(total_revenue, flipped_percent, department='HVAC', excused_hours=0):
    """
    Determine commission rate based on total revenue, flipped percentage, department, and excused hours.
    For every 8 hours excused, reduce thresholds by 20%.
    
    Args:
        total_revenue (float): Total revenue amount
        flipped_percent (float): Percentage of flipped sales
        department (str): Department name ('HVAC', 'Electric', or 'Plumbing')
        excused_hours (int): Number of excused hours
    """
    logger.debug(f"\nCalculating commission rate for:")
    logger.debug(f"Total Revenue: ${total_revenue:,.2f}")
    logger.debug(f"Initial Flipped %: {flipped_percent}")
    logger.debug(f"Department: {department}")
    logger.debug(f"Excused Hours: {excused_hours}")
    
    flipped_percent = custom_round_percentage(flipped_percent)
    logger.debug(f"Rounded Flipped %: {flipped_percent}")
    
    hvac_thresholds = {
        0: [7000, 8000, 9000, 10000],
        10: [7500, 8500, 9500, 11000],
        20: [8000, 9000, 10000, 11000],
        30: [9000, 10000, 11000, 12000],
        40: [10000, 11000, 12000, 13000],
        50: [12000, 14000, 16000, 18000],
        60: [14000, 16000, 18000, 20000],
        70: [15000, 17000, 19000, 21000],
        80: [16500, 18500, 20500, 23000],
        90: [18500, 20500, 23500, 26000],
        100: [22000, 24000, 26000, 29000]
    }
    
    plumbing_electrical_thresholds = {
        0: [7000, 8000, 9000, 10000],
        10: [7500, 8500, 9500, 11000],
        20: [8000, 9000, 10000, 11000],
        30: [9000, 10000, 11000, 12000],
        40: [10000, 11000, 12000, 13000],
        50: [11500, 12500, 14000, 15000],
        60: [13000, 14000, 15500, 17000],
        70: [14500, 15500, 17000, 18000],
        80: [15500, 16500, 18000, 19000],
        90: [17000, 18500, 19500, 21000],
        100: [18000, 20000, 22000, 24000]
    }
    
    # Use appropriate threshold table based on department
    if department == 'Electric' or department == 'Plumbing':
        thresholds = plumbing_electrical_thresholds
    else:
        thresholds = hvac_thresholds
        
    tier_thresholds = thresholds[flipped_percent].copy()  # Create a copy to avoid modifying the original
    
    # Calculate threshold reduction based on excused hours
    # Each 8 hours reduces thresholds by 20%
    reduction_factor = max(0, 1 - (0.20 * (excused_hours / 8)))
    logger.debug(f"Reduction factor from {excused_hours} excused hours: {reduction_factor}")
    
    # Apply reduction to thresholds
    adjusted_thresholds = [threshold * reduction_factor for threshold in tier_thresholds]
    logger.debug(f"Original thresholds for {flipped_percent}% flipped: {tier_thresholds}")
    logger.debug(f"Adjusted thresholds after {excused_hours} excused hours: {adjusted_thresholds}")
    
    if total_revenue >= adjusted_thresholds[3]:
        return 0.05
    elif total_revenue >= adjusted_thresholds[2]:
        return 0.04
    elif total_revenue >= adjusted_thresholds[1]:
        return 0.03
    elif total_revenue >= adjusted_thresholds[0]:
        return 0.02
    return 0

def autofit_columns(worksheet):
    """Autofit column widths in the worksheet."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    logger.info("Starting commission calculations")

    # Define the list of technicians to track
    technicians_to_track = [
        "Andrew Wycoff", "Andy Ventura", "Artie Straniti", "Brett Allen", "Carter Bruce",
        "Chris Smith", "Cody Hurlburt", "David Forney", "David Franklin", "David Knox",
        "Ethan Ficklin", "Gilberto Corvetto", "Glenn Griffin", "Hunter Stanley", 
        "Jacob Simpson", "Jake West", "Jason Kerns", "Josue Rodriguez", "Justin Barron", 
        "Kevin Stanley", "Pablo Silvas", "Patrick Bowerman", "Robert McGhee", "Ronnie Bland", 
        "Sean Lynders", "Shawn Hollingsworth", "Stephen Starner", "Thomas Ruiz-Lizama", 
        "Thomas Shawaryn", "Tim Kulesza", "WT Settle", "Will Winfree"
    ]

    # Load data file
    file_path = r'C:\Users\abatlouni\Downloads\combined_data.xlsx'
    
    logger.info(f"Loading data from {file_path}")
    data = pd.read_excel(file_path, sheet_name='Sheet1')

    # Filter data for specified technicians
    logger.info("Filtering for specified technicians")
    data = data[data['Primary Technician'].isin(technicians_to_track)]

    # Create the pivot table
    logger.info("Creating pivot table")
    pivot_table_revenue = data.pivot_table(
        index='Primary Technician',
        columns='Business Unit',
        values='Jobs Total Revenue',
        aggfunc='sum',
        fill_value=0
    )

    # Calculate the 'Total' column
    logger.info("Calculating total revenue")
    pivot_table_revenue.insert(0, 'Total', pivot_table_revenue.sum(axis=1))

    # Calculate 'Sold By Total'
    logger.info("Calculating sold by totals")
    sold_by_totals = (
        data[(data['Sold By'] != data['Primary Technician']) & 
             (data['Sold By'].isin(technicians_to_track))]
        .groupby('Sold By')['Jobs Total Revenue']
        .sum()
    )

    pivot_table_revenue.insert(1, 'Sold By Total', pivot_table_revenue.index.map(sold_by_totals).fillna(0))

    # Calculate 'Flipped %'
    logger.info("Calculating flipped percentages")
    flipped_percentage = (
        pivot_table_revenue['Sold By Total'] / (pivot_table_revenue['Total'] + pivot_table_revenue['Sold By Total']) * 100
    ).fillna(0).apply(custom_round_percentage)

    pivot_table_revenue.insert(2, 'Flipped %', flipped_percentage)

    # Get excused hours and add to pivot table
    logger.info("Getting excused hours from time off data")
    excused_hours = get_excused_hours(file_path)
    pivot_table_revenue.insert(3, 'Excused Hours', pivot_table_revenue.index.map(lambda x: excused_hours.get(str(x).strip(), 0)))

    # Calculate Commission % using department-specific thresholds
    logger.info("Calculating commission rates")
    
    # Add Department based on technician mapping
    logger.info("Adding department mapping")
    def get_department(tech_name):
        # Find numeric part in tech_name from Business Unit column
        tech_data = data[data['Primary Technician'] == tech_name]['Business Unit'].iloc[0]
        unit_number = ''
        for char in tech_data:
            if char.isdigit():
                unit_number += char
        
        if unit_number:
            num = int(unit_number)
            if 40 <= num <= 49:
                return 'Electric'
            elif 30 <= num <= 39:
                return 'Plumbing'
            elif 20 <= num <= 29:
                return 'HVAC'
        return 'Unknown'

    # Get departments for all technicians
    departments = {tech: get_department(tech) for tech in pivot_table_revenue.index}
    
    # Calculate commission rates using department information and excused hours
    commission_rates = pivot_table_revenue.apply(
        lambda row: get_commission_rate(
            row['Total'], 
            row['Flipped %'], 
            departments[row.name],
            row['Excused Hours']  # Pass excused hours to the function
        ),
        axis=1
    )
    pivot_table_revenue.insert(4, 'Commission %', commission_rates * 100)

    # Write to Excel
    output_path = r'C:\Users\abatlouni\Downloads\output.xlsx'
    logger.info(f"Writing results to {output_path}")
    
    # Create a new DataFrame with Department column in the desired position
    final_df = pd.DataFrame(index=pivot_table_revenue.index)
    final_df['Department'] = pd.Series(departments)
    
    # Combine with the rest of the data
    final_df = pd.concat([final_df, pivot_table_revenue], axis=1)
    
    # Write to Excel
    final_df.to_excel(output_path, sheet_name='Technician Revenue Totals')

    # Apply column autofitting
    wb = load_workbook(output_path)
    ws = wb['Technician Revenue Totals']
    autofit_columns(ws)
    wb.save(output_path)

    logger.info("Commission calculations completed")

if __name__ == "__main__":
    logger = setup_logging()
    main()